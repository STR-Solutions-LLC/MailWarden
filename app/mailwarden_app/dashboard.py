# (c) 2026 STR Solutions, LLC. All rights reserved.
"""
Dashboard — post-install UI. Seven tabs per §3.4.

Opens when MailWarden.app launches and ~/MailWarden/config/config.json exists
and parses as a valid config with at least one account.
"""
from __future__ import annotations

import csv
import json
import os
import subprocess
import sys
import tkinter as tk
import urllib.error
import urllib.request
import webbrowser
from datetime import datetime, timedelta
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

from . import config_io
from . import help_content
from . import paths
from . import smappservice_install
from . import startup_log
from . import theme
from . import validators


MODEL_CHOICES = [
    ("Haiku 4.5 (cheapest, default)", "claude-haiku-4-5-20251001"),
    ("Sonnet 4.6", "claude-sonnet-4-6"),
    ("Opus 4.7 (most capable)", "claude-opus-4-7"),
]


# =============================================================================
# Main window
# =============================================================================
class Dashboard(tk.Tk):
    def __init__(self):
        super().__init__()

        # Guard for the <Unmap> minimize backstop (Piece 2c) so it never fights
        # window teardown in _on_close. Set before any binding can fire.
        self._closing = False

        # PIECE 1 — Dashboard has NO Dock icon. Tk has just created the real
        # NSApplication in super().__init__(); this is the first moment NSApp()
        # is real, so demote it to an accessory (menu-bar-style, no Dock tile)
        # via PyObjC. LOUD: a future failure must never be silent. The open and
        # RAISE paths still call theme.bring_to_front (which activates +
        # focus_force + topmost-toggle) so the accessory app can still take
        # focus; bring_to_front itself logs that it ran.
        try:
            from AppKit import (  # type: ignore
                NSApp,
                NSApplicationActivationPolicyAccessory,
            )
            ns_app = NSApp()
            if ns_app is None:
                startup_log.step(
                    "dashboard: setActivationPolicy_(Accessory) SKIPPED — "
                    "NSApp() is None right after super().__init__() "
                    "(unexpected); a Dock icon may appear")
            else:
                ns_app.setActivationPolicy_(
                    NSApplicationActivationPolicyAccessory)
                startup_log.step(
                    "dashboard: setActivationPolicy_(Accessory) OK — "
                    "no Dock icon")
        except Exception as e:  # noqa: BLE001
            startup_log.step(
                f"dashboard: setActivationPolicy_(Accessory) FAILED "
                f"(Dock icon may appear): {type(e).__name__}: {e}")

        self.title(f"MailWarden  ·  {help_content.VERSION}")
        self.geometry("960x680")
        self.minsize(880, 600)
        theme.apply_theme(self)

        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        self.home_tab = HomeTab(self.notebook, self)
        self.accounts_tab = AccountsTab(self.notebook, self)
        self.lists_tab = ListsTab(self.notebook, self)
        self.signals_tab = SignalsTab(self.notebook, self)
        self.usage_tab = UsageTab(self.notebook, self)
        self.settings_tab = SettingsTab(self.notebook, self)
        self.diagnostics_tab = DiagnosticsTab(self.notebook, self)
        self.help_tab = HelpTab(self.notebook, self)

        for tab, title in [
            (self.home_tab, "Home"),
            (self.accounts_tab, "Accounts"),
            (self.lists_tab, "Whitelist / Blacklist"),
            (self.signals_tab, "Signal History"),
            (self.usage_tab, "API Usage"),
            (self.settings_tab, "Settings"),
            (self.diagnostics_tab, "Diagnostics"),
            (self.help_tab, "Help"),
        ]:
            self.notebook.add(tab, text=title)

        # Closing the window must QUIT this process (and free its Dock tile).
        # Without this, Tk keeps the process alive with a hidden root, leaving
        # an orphan Dock tile that can never reopen the window.
        self.protocol("WM_DELETE_WINDOW", self._on_close)

        # PIECE 2 — Dashboard CANNOT be minimized. An accessory app has no Dock
        # tile, so a minimized window would shrink to a Dock thumbnail that can
        # never be clicked to restore (no Dock tile to host it). Three layers,
        # weakest assumption last; the <Unmap> guard is the guaranteed backstop.
        #
        # (a) Disable the Cmd-M minimize shortcut — bind to a no-op that
        #     returns "break" so the event never reaches Tk's minimize handler.
        self.bind("<Command-m>", lambda _e: "break")
        self.bind("<Command-M>", lambda _e: "break")
        #
        # (b) Best-effort: remove the native minimize/collapse button via the
        #     unsupported MacWindowStyle API, keeping the close button and
        #     resizing. The exact atom string differs between system Tk and the
        #     bundled Tk, so try the documented forms and log which (if any)
        #     took. Never rely on this alone.
        minimize_style_result = "not applied"
        try:
            from tkinter import TclError as _TclError
            for _attrs in (
                "closeBox collapseBox resizable",   # explicit, sans miniaturize
                "document closeBox resizable",      # document w/o miniaturize
            ):
                try:
                    self.tk.call(
                        "::tk::unsupported::MacWindowStyle", "style",
                        self._w, "document", _attrs)
                    minimize_style_result = f"MacWindowStyle attrs='{_attrs}'"
                    break
                except _TclError:
                    continue
        except Exception as e:  # noqa: BLE001
            minimize_style_result = f"failed ({type(e).__name__}: {e})"
        #
        # (b2) Most reliable layer on this Tk/macOS build: clear the
        #     NSWindowStyleMaskMiniaturizable bit on the underlying NSWindow via
        #     PyObjC, which greys out (disables) the yellow minimize button so it
        #     cannot be clicked at all, while leaving close + resize intact. The
        #     NSWindow only exists once Tk has realized the window, so this runs
        #     after the first idle. Best-effort and logged; never relied on alone.
        self.after_idle(self._disable_native_minimize_button)
        #
        # (c) GUARANTEED BACKSTOP: if a minimize sneaks through (Cmd-M variant,
        #     window-menu item, or a Tk build that ignores the style above),
        #     <Unmap> fires and self.state() reports "iconic". As long as we are
        #     not closing, snap the window straight back open via after_idle so
        #     no Dock thumbnail can persist.
        self.bind("<Unmap>", self._on_unmap)
        startup_log.step(
            f"dashboard: minimize disabled — MacWindowStyle="
            f"{minimize_style_result}; Cmd-M bound; Unmap guard on")

        # Dock-tile reopen/restore: rely on Tk's OWN native delegate
        # (TKApplication already implements applicationShouldHandleReopen:
        # hasVisibleWindows: and the standard minimized-window un-minimize).
        # We log what Tk installed so on-device we can confirm it's present;
        # we do NOT replace it (the prior attempt did, which never fired AND
        # stripped Tk's native window management — see _log_dock_reopen_state).
        self._log_dock_reopen_state()

        # Cross-process RAISE channel: a loopback socket server, running on a
        # daemon thread, that other processes (the menu bar agent, the
        # Applications launch guard, and the menu bar's double-click reopen
        # handler) hit to un-minimize and raise THIS window. Only this process
        # can deiconify() a minimized window, so the socket is the one reliable
        # cross-process un-minimize. Started after
        # the window exists; torn down in _on_close (and in app_entrypoint's
        # finally) so no port file or thread outlives the process.
        from . import dashboard_ipc
        self._raise_server = dashboard_ipc.DashboardRaiseServer(
            self, theme.bring_to_front)
        self._raise_server.start()

        # Refresh whichever tab the user switches to so changes made by the
        # filter (it writes whitelist.json / blacklist.json / token_usage.json
        # in a separate process) show up the next time the user clicks the
        # tab. Before this binding, the filter could add 50 whitelist entries
        # and the Lists tab would still show the snapshot from Dashboard
        # launch until a full refresh_all() happened to fire.
        self.notebook.bind("<<NotebookTabChanged>>", self._on_tab_changed)

        self.after(60_000, self._periodic_refresh)

        # Silent backstop: auto-create the Train MailWarden folder on every
        # enabled account that is missing it. Runs once per Dashboard launch on
        # a background thread; logs INFO only when a folder is actually created.
        # This is idempotent — if the folder already exists, ensure_train_folder
        # returns immediately via LIST discovery.
        self.after(1000, self._autocreate_train_folders)

        # After the UI has settled, check each enabled account for the
        # "Train MailWarden" IMAP folder and offer to create it if it's
        # missing. Respects ui.prompt_missing_train_folder so users who
        # decline or have their own setup can silence it.
        self.after(2500, self._maybe_check_train_folders)

        # PIECE 5 (Dashboard side) — show the one-time "MailWarden is running"
        # message exactly once per update, then stamp the version. Deferred via
        # after() (NOT inline) so it cannot deadlock startup or block the window
        # coming to front. Whether the menu-bar auto-opened the Dashboard or the
        # user opened it manually, this is the single place that shows the
        # message and writes the stamp — guaranteeing once-per-version.
        self.after(800, self._maybe_show_welcome_once)

    def _on_unmap(self, _event=None):
        """GUARANTEED minimize backstop (Piece 2c). Fires on <Unmap>; if the
        window went iconic (minimized) and we are not tearing down, snap it
        back open on the next idle so no Dock thumbnail can persist."""
        if getattr(self, "_closing", False):
            return
        try:
            if self.state() == "iconic":
                self.after_idle(self.deiconify)
        except tk.TclError:
            pass

    def _disable_native_minimize_button(self):
        """Disable the macOS minimize (yellow) button by clearing
        NSWindowStyleMaskMiniaturizable on the Tk NSWindow (Piece 2b2).

        The MacWindowStyle atom strings are unreliable across Tk builds; the
        NSWindow styleMask is the authoritative source of which title-bar
        buttons are active. We locate our NSWindow by matching its title (set
        in __init__) among NSApp().windows(), then AND-out the miniaturizable
        bit. Close + resize bits are left untouched. Best-effort and logged;
        the <Unmap> backstop still guards anything that slips through."""
        try:
            from AppKit import NSApp  # type: ignore
            ns_app = NSApp()
            if ns_app is None:
                startup_log.step(
                    "dashboard: minimize-button disable SKIPPED — NSApp() None")
                return
            wanted = self.title()
            target = None
            for win in ns_app.windows():
                try:
                    if str(win.title()) == wanted:
                        target = win
                        break
                except Exception:  # noqa: BLE001
                    continue
            if target is None:
                # Fall back to the key/main window if the title lookup misses.
                target = ns_app.mainWindow() or ns_app.keyWindow()
            if target is None:
                startup_log.step(
                    "dashboard: minimize-button disable SKIPPED — no NSWindow "
                    "found (title match + key/main both empty)")
                return
            # NSWindowStyleMaskMiniaturizable == 1 << 2 (stable AppKit constant).
            miniaturizable = 1 << 2
            mask = int(target.styleMask())
            target.setStyleMask_(mask & ~miniaturizable)
            startup_log.step(
                "dashboard: minimize button DISABLED via NSWindow styleMask "
                f"(mask {mask} -> {mask & ~miniaturizable})")
        except Exception as e:  # noqa: BLE001
            startup_log.step(
                f"dashboard: minimize-button disable FAILED (non-fatal, "
                f"Unmap backstop still active): {type(e).__name__}: {e}")

    def _maybe_show_welcome_once(self):
        """Show the one-time welcome message + stamp the version (Piece 5).
        Best-effort, logged, non-fatal — a stamp read/write hiccup must never
        crash the Dashboard."""
        try:
            from . import bootstrap
            current = bootstrap.BUNDLED_FILTER_VERSION
            stamp = config_io.load_json(paths.WELCOME_SHOWN_PATH, {})
            shown_for = stamp.get("version") if isinstance(stamp, dict) else None
            if shown_for == current:
                return
            messagebox.showinfo(
                "MailWarden is running",
                "You can close this window anytime; MailWarden keeps filtering "
                "your mail in the background. Open the dashboard in the future "
                "by clicking the MailWarden menu icon.",
                parent=self,
            )
            try:
                config_io.save_json_atomic(
                    paths.WELCOME_SHOWN_PATH,
                    {"version": current, "shown_at": config_io.now_iso()},
                )
                startup_log.step(
                    f"dashboard: welcome message shown and stamped to {current}")
            except Exception as e:  # noqa: BLE001
                startup_log.step(
                    f"dashboard: welcome stamp write FAILED (non-fatal): "
                    f"{type(e).__name__}: {e}")
        except Exception as e:  # noqa: BLE001
            try:
                startup_log.step(
                    f"dashboard: welcome message check FAILED (non-fatal): "
                    f"{type(e).__name__}: {e}")
            except Exception:
                pass

    def _on_close(self):
        """Closing the window quits the process. Tearing down the Tk mainloop
        lets run()'s mainloop() return; app_entrypoint's finally then removes
        the dashboard pidfile, freeing this process's Dock tile. Without this
        the process would linger with a hidden root and leak its Dock icon."""
        # PIECE 3 — mark teardown so the <Unmap> minimize guard (Piece 2c)
        # bails out instead of trying to deiconify a window we're destroying.
        self._closing = True
        # Tear down the RAISE socket + remove its port file before quitting, so
        # no stale dashboard.port lingers and the daemon thread's listening
        # socket is closed cleanly (it can't keep the process alive, but this
        # avoids a brief window where a stale port answers a connect).
        try:
            if getattr(self, "_raise_server", None) is not None:
                self._raise_server.stop()
        except Exception:
            pass
        try:
            self.quit()
        except tk.TclError:
            pass
        try:
            self.destroy()
        except tk.TclError:
            pass

    def _log_dock_reopen_state(self):
        """Instrument (do NOT replace) the Tk NSApplication delegate.

        On-device evidence (beta.6-test logs) proved the prior approach —
        installing our OWN bare NSObject delegate via NSApp().setDelegate_()
        and waiting for applicationShouldHandleReopen:hasVisibleWindows: to
        fire — NEVER fired ("dock reopen fired" never appeared). Investigation
        on the build machine found why: under Tk, NSApp() is a TKApplication
        whose delegate is TKApplication itself, and TKApplication ALREADY
        implements applicationShouldHandleReopen:hasVisibleWindows: plus the
        standard minimized-window restore. Replacing that delegate with a bare
        object both (a) failed to receive a minimized-window-thumbnail click
        (WindowServer un-minimizes the thumbnail directly; it does not route a
        reopen Apple Event to our bare delegate) and (b) STRIPPED Tk's native
        window management. So the supported, reliable un-minimize paths are:
          - menu bar "Open Dashboard"  -> reopen_or_spawn() -> RAISE socket
          - double-click MailWarden.app -> reopen handler in the menu bar
            process (menu_bar._install_app_reopen_handler) -> reopen_or_spawn()
          - the macOS-native minimized-window thumbnail click, handled by Tk
            itself (left intact by NOT replacing the delegate).
        This method only LOGS what delegate Tk installed so we can confirm the
        native handler is present on-device; it changes no behavior."""
        try:
            from AppKit import NSApp  # type: ignore
            ns_app = NSApp()
            if ns_app is None:
                startup_log.step(
                    "dashboard: NSApp() is None at startup; cannot inspect "
                    "Tk delegate (Dock reopen relies on menu/Applications)")
                return
            delegate = ns_app.delegate()
            cls = type(delegate).__name__ if delegate is not None else "None"
            responds = (
                delegate is not None
                and delegate.respondsToSelector_(
                    "applicationShouldHandleReopen:hasVisibleWindows:"))
            startup_log.step(
                f"dashboard: Tk NSApplication delegate={cls}, "
                f"native reopen handler present={responds} "
                f"(not replaced — Tk owns minimized-window restore)")
        except Exception as e:  # noqa: BLE001
            try:
                startup_log.step(
                    f"dashboard: could not inspect Tk delegate (non-fatal): "
                    f"{type(e).__name__}: {e}")
            except Exception:
                pass

    def refresh_all(self):
        self.home_tab.refresh()
        self.accounts_tab.refresh()
        self.lists_tab.refresh()
        self.signals_tab.refresh()
        self.usage_tab.refresh()
        self.settings_tab.refresh()
        self.diagnostics_tab.refresh()

    def _periodic_refresh(self):
        # The filter runs in a separate process every few minutes and writes
        # whitelist.json / blacklist.json / signals / decisions behind the
        # Dashboard's back. Periodic refresh keeps the visible tabs in sync
        # without forcing the user to close and reopen the window.
        try:
            self.home_tab.refresh()
            self.usage_tab.refresh()
            self.lists_tab.refresh()
            self.signals_tab.refresh()
        finally:
            self.after(60_000, self._periodic_refresh)

    # ---- Train MailWarden folder auto-create + prompt check ----

    TRAIN_FOLDER_NAME = "Train MailWarden"

    def _autocreate_train_folders(self):
        """Silent backstop: attempt to create the Train MailWarden folder on
        every enabled account that doesn't have it yet. Fires once per
        Dashboard launch from a background thread. No dialog, no prompt — logs
        INFO only when a creation actually occurs. Any IMAP failure is caught
        and silently discarded so Dashboard startup is never affected.

        Thread pattern follows _on_rebuild_schedulers: worker thread, polled
        via self.after(150, ...).
        """
        import threading
        config = config_io.load_config()
        accounts = [a for a in (config.get("accounts") or [])
                    if a.get("enabled", True)]
        if not accounts:
            return

        import logging
        logger = logging.getLogger("mailwarden.dashboard")

        def _worker():
            import imaplib
            for account in accounts:
                label = account.get("name") or account.get("username", "?")
                try:
                    conn = imaplib.IMAP4_SSL(
                        account["imap_host"],
                        int(account.get("imap_port", 993)),
                        timeout=10,
                    )
                    try:
                        conn.login(account["username"], account["password"])
                        ok, msg = self._ensure_train_folder_bg(conn)
                        if ok and "already exists" not in msg:
                            logger.info(
                                f"Created Train MailWarden folder on {label!r}"
                                f" ({msg})")
                        # If folder already existed, skip silently.
                    finally:
                        try:
                            conn.logout()
                        except Exception:
                            pass
                except Exception:
                    # Network / auth failures are silently skipped —
                    # this is a best-effort backstop, not a critical path.
                    pass

        t = threading.Thread(target=_worker, daemon=True)
        t.start()

    def _ensure_train_folder_bg(self, conn) -> tuple[bool, str]:
        """Idempotently ensure the Train MailWarden folder exists on conn.

        Mirrors ensure_train_folder() in spam_filter.py. Kept here (rather
        than calling spam_filter directly) because the Dashboard is in the app
        bundle, not the installed filter package; importing spam_filter from
        the app context would require sys.path manipulation.

        Step 1: wildcard LIST — returns quickly if folder already exists.
        Step 2: CREATE "Train MailWarden" (quoted; [ALREADYEXISTS] = success).
        Step 3: CREATE "INBOX.Train MailWarden" fallback for Cyrus/Bluehost.
        Step 4: both failed — return (False, reason). Does NOT raise.
        """
        target = self.TRAIN_FOLDER_NAME.lower()
        try:
            rc_list, items = conn.list('""', '"*"')
        except Exception:
            rc_list, items = "NO", []

        if rc_list == "OK" and items:
            for raw in items:
                if not raw:
                    continue
                line = raw.decode("utf-8", errors="replace") if isinstance(
                    raw, (bytes, bytearray)) else str(raw)
                if '"' in line:
                    parts = line.rsplit('"', 2)
                    if len(parts) >= 2:
                        name = parts[-2]
                        if target in name.lower():
                            return (True, f"{name} already exists")
                else:
                    tail = line.rsplit(None, 1)[-1] if line.split() else ""
                    if target in tail.lower():
                        return (True, f"{tail} already exists")

        def _decode(data) -> str:
            return b" ".join(
                x for x in (data or []) if x
            ).decode("utf-8", errors="replace")

        first_err = ""
        try:
            rc, data = conn.create(f'"{self.TRAIN_FOLDER_NAME}"')
            if rc == "OK":
                return (True, self.TRAIN_FOLDER_NAME)
            detail = _decode(data).lower()
            if "alreadyexists" in detail or "already exists" in detail:
                return (True, self.TRAIN_FOLDER_NAME)
            first_err = detail
        except Exception as e:
            first_err = str(e)

        inbox_name = f"INBOX.{self.TRAIN_FOLDER_NAME}"
        try:
            rc2, data2 = conn.create(f'"{inbox_name}"')
            if rc2 == "OK":
                return (True, inbox_name)
            detail2 = _decode(data2).lower()
            if "alreadyexists" in detail2 or "already exists" in detail2:
                return (True, inbox_name)
            detail2_raw = detail2
        except Exception as e:
            detail2_raw = str(e)

        return (False, (f"CREATE failed: top-level={first_err!r}; "
                        f"INBOX. fallback={detail2_raw!r}"))

    def _maybe_check_train_folders(self):
        config = config_io.load_config()
        if not config.get("ui", {}).get("prompt_missing_train_folder", True):
            return
        # Run IMAP probe in a background thread so Dashboard stays responsive
        # on a slow link.
        import threading
        threading.Thread(target=self._check_train_folders_bg,
                          args=(config,), daemon=True).start()

    def _check_train_folders_bg(self, config: dict):
        import imaplib
        missing = []
        for account in config.get("accounts", []) or []:
            if not account.get("enabled", True):
                continue
            try:
                conn = imaplib.IMAP4_SSL(account["imap_host"],
                                          int(account.get("imap_port", 993)),
                                          timeout=10)
                try:
                    conn.login(account["username"], account["password"])
                    # Use the wildcard helper so we find the folder whether
                    # it lives at top-level ("Train MailWarden") or nested
                    # under INBOX ("INBOX.Train MailWarden") on Cyrus/Bluehost.
                    # A bare LIST with the exact name misses the nested case
                    # and nags users whose folder already exists on v1.5.9+.
                    exists = self._find_train_folder(conn) is not None
                    if not exists:
                        missing.append(account)
                finally:
                    try:
                        conn.logout()
                    except Exception:
                        pass
            except Exception:
                # Network hiccups shouldn't pester the user on launch;
                # next launch will re-check.
                pass
        if missing:
            self.after(0, lambda: self._prompt_create_train_folder(missing))

    def _prompt_create_train_folder(self, accounts: list[dict]):
        names = "\n  • ".join(
            a.get("name") or a.get("username", "?") for a in accounts)
        if not messagebox.askyesno(
            "Train MailWarden folder",
            f"The following accounts don't have a \"Train MailWarden\" "
            f"folder yet:\n\n  • {names}\n\n"
            f"This folder lets you train the spam filter by dragging spam "
            f"examples into it — the primary way to teach MailWarden. "
            f"Create it in each account now? (You can turn off this prompt "
            f"in Settings.)"):
            return
        import threading
        threading.Thread(target=self._create_train_folders_bg,
                          args=(accounts,), daemon=True).start()

    def _find_train_folder(self, conn) -> str | None:
        """Return the visible mailbox name of the Train MailWarden
        folder on this connection, or None if it isn't visible at all.

        Wildcard-LISTs the entire mailbox tree and matches case-
        insensitively on the folder name — covers top-level on AOL/
        Gmail/iCloud, INBOX.Train MailWarden on Cyrus, and INBOX/Train
        MailWarden on slash-delimiter servers, all in one pass.
        """
        target = self.TRAIN_FOLDER_NAME.lower()
        try:
            rc, items = conn.list('""', '"*"')
        except Exception:
            return None
        if rc != "OK" or not items:
            return None
        for raw in items:
            if not raw:
                continue
            line = raw.decode("utf-8", errors="replace") if isinstance(
                raw, (bytes, bytearray)) else str(raw)
            # IMAP LIST line: (\HasNoChildren) "/" "INBOX.Train MailWarden"
            # The mailbox name is the last quoted token. Split from the
            # right so the path with possible spaces stays intact.
            if '"' in line:
                # Take the last quoted segment.
                parts = line.rsplit('"', 2)
                if len(parts) >= 2:
                    name = parts[-2]
                    if target in name.lower():
                        return name
            else:
                # No quotes — the name is the last whitespace-separated
                # token (rare; happens for ASCII-only single-word names).
                tail = line.rsplit(None, 1)[-1] if line.split() else ""
                if target in tail.lower():
                    return tail
        return None

    def _create_train_folders_bg(self, accounts: list[dict]):
        import imaplib
        results = []
        for a in accounts:
            label = a.get("name") or a.get("username", "?")
            try:
                conn = imaplib.IMAP4_SSL(a["imap_host"],
                                          int(a.get("imap_port", 993)),
                                          timeout=10)
                try:
                    conn.login(a["username"], a["password"])
                    # Quote the mailbox name. Python's imaplib does NOT
                    # auto-quote, and IMAP servers parse a bare folder
                    # name with a space as two arguments — AOL rejects
                    # with [CLIENTBUG] "Additional arguments found",
                    # Bluehost/Cyrus returns NO. Quoting fixes both.
                    quoted = f'"{self.TRAIN_FOLDER_NAME}"'
                    # If the server hints that folders need an INBOX. prefix
                    # (Bluehost/Cyrus: "Mailbox name should probably be
                    # prefixed with: INBOX."), skip the bare attempt and go
                    # straight to the INBOX. form to keep logs clean.
                    rc, data = conn.create(quoted)
                    if rc != "OK":
                        data_str = b" ".join(
                            x for x in (data or []) if x
                        ).decode("utf-8", errors="replace").lower()
                        # [ALREADYEXISTS] counts as success — folder already
                        # exists, user's intent is satisfied.
                        if "alreadyexists" in data_str or "already exists" in data_str:
                            rc = "OK"
                        else:
                            # Fallback for Cyrus-style personal namespaces.
                            # If Bluehost hinted "should probably be prefixed
                            # with: INBOX." go straight to INBOX. form without
                            # logging a spurious failure for the bare attempt.
                            rc2, data2 = conn.create(
                                f'"INBOX.{self.TRAIN_FOLDER_NAME}"')
                            data2_str = b" ".join(
                                x for x in (data2 or []) if x
                            ).decode("utf-8", errors="replace").lower()
                            if rc2 == "OK" or "alreadyexists" in data2_str or "already exists" in data2_str:
                                rc = "OK"
                                data = data2
                            else:
                                data = (data or []) + [b"; INBOX. fallback: "] \
                                    + (data2 or [])
                    # Verify CREATE actually produced a visible folder.
                    # Some servers return OK but stash the mailbox in a
                    # namespace the user's mail client doesn't display
                    # — we hit this on Dad's AOL account with v1.5.9.
                    # The only reliable test is to LIST and see if it's
                    # actually there.
                    visible_as = self._find_train_folder(conn)
                    if rc == "OK" and visible_as:
                        # Tell the user where the folder lives if it's
                        # nested rather than top-level, so they don't
                        # hunt for an invisible folder.
                        if visible_as != self.TRAIN_FOLDER_NAME:
                            results.append(
                                (label, True,
                                 f"created as \"{visible_as}\""))
                        else:
                            results.append((label, True, ""))
                    elif rc == "OK" and not visible_as:
                        results.append(
                            (label, False,
                             "server reported CREATE OK but the folder "
                             "is not visible via LIST — your mail "
                             "provider may have rejected it silently."))
                    else:
                        detail = b" ".join(
                            x for x in (data or []) if x).decode(
                                "utf-8", errors="replace")
                        results.append(
                            (label, False,
                             f"IMAP CREATE rc={rc} {detail}".strip()))
                finally:
                    try:
                        conn.logout()
                    except Exception:
                        pass
            except Exception as e:
                results.append((label, False, str(e)))
        self.after(0, lambda: self._report_train_folder_results(results))

    def _report_train_folder_results(self, results: list[tuple]):
        lines = []
        any_failed = False
        for label, ok, detail in results:
            if ok:
                # detail is empty for the common top-level case, or a
                # descriptive string when the folder ended up nested.
                if detail:
                    lines.append(f"  ✓ {label}: {detail}")
                else:
                    lines.append(f"  ✓ {label}: created")
            else:
                any_failed = True
                lines.append(f"  ✗ {label}: {detail}")
        if any_failed:
            messagebox.showwarning(
                "Train MailWarden folder",
                "Some folders could not be created. You can retry from "
                "Settings.\n\n" + "\n".join(lines))
        else:
            messagebox.showinfo(
                "Train MailWarden folder",
                "Train MailWarden folder created successfully.\n\n" +
                "\n".join(lines) +
                "\n\nTo train the filter, drag any spam email into this "
                "folder. MailWarden will email you a refinement proposal "
                "within 15 minutes.")

    def _on_tab_changed(self, _evt=None):
        """Refresh the tab the user just switched to. Each tab's refresh()
        is cheap (reads 1–3 JSON files, max); no point being clever."""
        try:
            idx = self.notebook.index(self.notebook.select())
        except Exception:
            return
        tabs = [self.home_tab, self.accounts_tab, self.lists_tab,
                self.signals_tab, self.usage_tab, self.settings_tab,
                self.diagnostics_tab, self.help_tab]
        if 0 <= idx < len(tabs):
            tab = tabs[idx]
            try:
                if hasattr(tab, "refresh"):
                    tab.refresh()
            except Exception:
                # A refresh error should never prevent the user from seeing
                # the tab content; swallow and continue.
                pass


# =============================================================================
# HOME
# =============================================================================
class HomeTab(ttk.Frame):
    def __init__(self, parent, app: Dashboard):
        super().__init__(parent)
        self.app = app
        self._scroll = _ScrollableTab(self)
        self._scroll.pack(fill=tk.BOTH, expand=True)
        self._f = ttk.Frame(self._scroll.body, padding=(16, 12))
        self._f.pack(fill=tk.BOTH, expand=True)
        # after() id for the post-Run-Now last-run poll, so a pending poll can
        # be cancelled if this tab is torn down mid-poll.
        self._run_now_poll_after: str | None = None
        self.bind("<Destroy>", self._cancel_run_now_poll, add="+")
        self._build()
        self.refresh()

    def _build(self):
        ttk.Label(self._f, text="MailWarden", style="Heading.TLabel").pack(
            anchor=tk.W, pady=(0, 4))
        ttk.Label(self._f, style="Muted.TLabel",
                  text="Live status of your AI-powered spam filter").pack(
            anchor=tk.W, pady=(0, 18))

        # --- Account status card ---
        status_card = ttk.LabelFrame(self._f, text="Status",
                                      padding=(14, 12))
        status_card.pack(fill=tk.X, pady=(0, 14))
        self._accounts_frame = ttk.Frame(status_card)
        self._accounts_frame.pack(fill=tk.X)

        # --- Filter activity card — two-column layout ---
        # Left: summary stats + spam counters. Right: per-account breakdown.
        stats_outer = ttk.LabelFrame(self._f,
                                      text="Filter activity (last 24 h / this week / lifetime)",
                                      padding=(14, 12))
        stats_outer.pack(fill=tk.X, pady=(0, 14))
        stats_outer.columnconfigure(0, weight=3)
        stats_outer.columnconfigure(1, weight=2)

        # Left column
        stats_left = ttk.Frame(stats_outer)
        stats_left.grid(row=0, column=0, sticky=tk.NSEW)

        self._stats_labels: dict[str, ttk.Label] = {}
        for i, label in enumerate(("Emails evaluated today:",
                                    "Emails evaluated this week:",
                                    "Emails evaluated lifetime:",
                                    "Spam caught today:",
                                    "Spam caught — lifetime:",
                                    "Last filter run:")):
            ttk.Label(stats_left, text=label, style="Muted.TLabel").grid(
                row=i, column=0, sticky=tk.W, pady=3)
            v = ttk.Label(stats_left, text="—", style="Subheading.TLabel")
            v.grid(row=i, column=1, sticky=tk.W, padx=(18, 0), pady=3)
            self._stats_labels[label] = v

        # Right column — per-account breakdown
        ttk.Separator(stats_outer, orient=tk.VERTICAL).grid(
            row=0, column=1, sticky=tk.NS, padx=(18, 0))
        self._per_account_frame = ttk.Frame(stats_outer, padding=(14, 0))
        self._per_account_frame.grid(row=0, column=2, sticky=tk.NSEW, padx=(0, 0))
        ttk.Label(self._per_account_frame,
                  text="Evaluated today — per account",
                  style="Muted.TLabel").pack(anchor=tk.W, pady=(0, 6))

        # --- Mode + actions card ---
        mode = ttk.LabelFrame(self._f, text="Operating mode", padding=(14, 12))
        mode.pack(fill=tk.X, pady=(0, 14))
        self._dry_run_var = tk.BooleanVar()
        self._dry_run_chk = ttk.Checkbutton(
            mode, text="Dry run — classify but do not move any mail",
            variable=self._dry_run_var,
            command=self._on_toggle_dry_run)
        self._dry_run_chk.pack(anchor=tk.W)
        ttk.Label(mode, style="Muted.TLabel",
                  text="Turn off dry run once you've reviewed a few "
                       "classifications and are confident MailWarden is "
                       "acting sensibly on your mail.").pack(anchor=tk.W, pady=(6, 0))

        # --- Action buttons ---
        actions = ttk.Frame(self._f)
        actions.pack(fill=tk.X, pady=(0, 4))
        ttk.Button(actions, text="Run Now", style="Primary.TButton",
                    command=self._on_run_now).pack(side=tk.LEFT)
        ttk.Button(actions, text="Reset cache and re-scan inbox",
                    command=self._on_reset_cache).pack(side=tk.LEFT, padx=(10, 0))
        ttk.Button(actions, text="View Recent Decisions",
                    command=self._on_view_decisions).pack(side=tk.LEFT, padx=(10, 0))
        ttk.Button(actions, text="Open Log File",
                    command=self._on_open_log).pack(side=tk.LEFT, padx=(10, 0))

        # Update banner (hidden unless update available)
        self._update_banner = ttk.Frame(self._f, padding=(8, 8), relief=tk.GROOVE)
        self._update_banner_label = ttk.Label(self._update_banner,
                                                 text="", foreground="#226")
        self._update_banner_label.pack(side=tk.LEFT)
        ttk.Button(self._update_banner, text="View Release Notes",
                   command=self._on_open_releases).pack(side=tk.RIGHT)

    def refresh(self):
        config = config_io.load_config()

        for w in self._accounts_frame.winfo_children():
            w.destroy()
        last_run = _last_filter_run()
        state, long_line = _determine_health(last_run)
        shape = {"green": "●", "yellow": "◐", "red": "○"}[state]
        status_style = {"green": "Green.TLabel", "yellow": "Yellow.TLabel",
                         "red": "Red.TLabel"}[state]
        header = ttk.Label(self._accounts_frame,
                           text=f"{shape}  {long_line}",
                           style=status_style)
        header.pack(anchor=tk.W)

        for a in config.get("accounts", []):
            row = ttk.Frame(self._accounts_frame)
            row.pack(fill=tk.X, pady=2)
            badge = "●" if a.get("enabled") else "○"
            ttk.Label(row,
                      text=f"{badge}  {a.get('name','(unnamed)')}  —  {a.get('username','(no address)')}",
                      foreground=("#1a7f37" if a.get("enabled") else "#666")).pack(side=tk.LEFT)

        # Stats from decisions.log
        today, week, life = _decision_counts()
        spam_today, spam_life = _spam_killed_counts()
        self._stats_labels["Emails evaluated today:"].config(text=f"{today}")
        self._stats_labels["Emails evaluated this week:"].config(text=f"{week}")
        self._stats_labels["Emails evaluated lifetime:"].config(text=f"{life}")
        self._stats_labels["Spam caught today:"].config(text=f"{spam_today}")
        self._stats_labels["Spam caught — lifetime:"].config(text=f"{spam_life}")
        self._stats_labels["Last filter run:"].config(
            text=(last_run.strftime("%Y-%m-%d %H:%M") if last_run else "never"))

        # Per-account today breakdown — right column of the filter activity frame.
        # Destroy all children EXCEPT the header label (first child added in _build).
        children = self._per_account_frame.winfo_children()
        for w in children[1:]:
            w.destroy()
        accounts = config.get("accounts", [])
        per_acct = _decision_counts_by_account()
        # Use monospace font so the dot-leader columns align
        mono = ("Menlo", 11)
        if accounts:
            for a in accounts:
                acct_name = a.get("name") or a.get("username", "?")
                count = per_acct.get(acct_name, 0)
                # Dot-leader: pad name, fill dots, right-align count
                dots = "." * max(4, 30 - len(acct_name) - len(str(count)))
                row = ttk.Frame(self._per_account_frame)
                row.pack(fill=tk.X, pady=1)
                ttk.Label(row,
                          text=f"  \u2022 {acct_name}  {dots}  {count}",
                          font=mono).pack(anchor=tk.W)
            # Separator + total
            ttk.Separator(self._per_account_frame,
                           orient=tk.HORIZONTAL).pack(fill=tk.X, pady=(4, 2))
            total_dots = "." * max(4, 30 - len("Total") - len(str(today)))
            ttk.Label(self._per_account_frame,
                      text=f"  Total  {total_dots}  {today}",
                      font=mono).pack(anchor=tk.W, pady=(0, 2))
        else:
            ttk.Label(self._per_account_frame,
                      text="No accounts configured.",
                      style="Muted.TLabel").pack(anchor=tk.W)

        # Dry run
        self._dry_run_var.set(bool(config.get("filter", {}).get("dry_run", True)))

        # Update check (throttled to once per 14 days)
        self._maybe_show_update_banner()

    def _on_toggle_dry_run(self):
        new_val = bool(self._dry_run_var.get())
        if not new_val:
            if not messagebox.askyesno(
                "Turn off dry run?",
                "MailWarden will start actually moving spam into your junk folder. "
                "Make sure you've reviewed a few recent decisions in the log "
                "before continuing.\n\nProceed?",
            ):
                self._dry_run_var.set(True)
                return
        config = config_io.load_config()
        config.setdefault("filter", {})["dry_run"] = new_val
        config_io.save_config(config)
        self.refresh()

    def _on_run_now(self):
        if _lock_active():
            messagebox.showinfo("Filter is already running",
                                 "Wait a moment — the filter is already running.")
            return
        filter_script = paths.SRC_DIR / "spam_filter.py"
        if not filter_script.exists():
            messagebox.showerror("Missing filter",
                                 f"Filter script missing: {filter_script}")
            return
        app_bundle = Path("/Applications/MailWarden.app")
        bundled_python = app_bundle / "Contents" / "MacOS" / "python"
        launcher = app_bundle / "Contents" / "Resources" / "launcher.py"
        if not bundled_python.exists() or not launcher.exists():
            messagebox.showerror(
                "App bundle missing",
                f"Expected {app_bundle}. Reinstall MailWarden.")
            return
        # Invoke the bundled Python directly on launcher.py. This bypasses
        # LaunchServices and NSApplicationMain completely — no new app
        # instance bounces in the Dock, no single-bundle-instance collision,
        # no Launch Error dialog. launcher.py still sets sys.path, UTF-8
        # open patch, and SSL_CERT_FILE so the filter imports cleanly.
        env = dict(os.environ)
        env["PYTHONUTF8"] = "1"
        env["LC_ALL"] = "en_US.UTF-8"
        env["LANG"] = "en_US.UTF-8"
        try:
            subprocess.Popen(
                # --force bypasses the interval gate in spam_filter.run_filter
                # so manual Run Now always runs immediately, even if the
                # scheduled agent ran moments ago. The launchd agent invokes
                # the same launcher WITHOUT --force, so scheduled wakes stay
                # gated by filter.interval_minutes.
                [str(bundled_python), str(launcher), "--run-filter", "--force"],
                cwd=str(paths.MAILWARDEN_ROOT),
                env=env,
                stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                start_new_session=True,
            )
        except OSError as e:
            messagebox.showerror("Could not launch", str(e))
            return
        # The filter runs async in a subprocess and only advances FILTER_LOG's
        # mtime (what _last_filter_run reads) once it actually executes, which
        # is after this method returns. A one-shot refresh here would still show
        # the old "Last Run". Instead, snapshot the current run time and poll
        # the Home tab refresh every 2s for ~30s, stopping as soon as the run
        # timestamp advances. The Dashboard's 60s _periodic_refresh remains the
        # backstop so the label is never permanently stale.
        baseline = _last_filter_run()
        self._start_run_now_poll(baseline, deadline_ticks=15)
        messagebox.showinfo("Running",
                             "Filter started. Come back in a minute to see updated stats.")

    def _start_run_now_poll(self, baseline, deadline_ticks: int):
        """Poll _last_filter_run() every 2s up to deadline_ticks times, calling
        refresh() each tick, and stop early once the run timestamp advances past
        baseline. Guarded against teardown: any pending poll is cancelled on
        <Destroy>, and each tick bails if the widget no longer exists."""
        self._cancel_run_now_poll()

        def _tick(remaining: int):
            self._run_now_poll_after = None
            try:
                if not self.winfo_exists():
                    return
            except tk.TclError:
                return
            try:
                self.refresh()
            except Exception:
                pass
            current = _last_filter_run()
            advanced = current is not None and (
                baseline is None or current > baseline)
            if advanced or remaining <= 0:
                return
            try:
                self._run_now_poll_after = self.after(
                    2000, lambda: _tick(remaining - 1))
            except tk.TclError:
                self._run_now_poll_after = None

        try:
            self._run_now_poll_after = self.after(
                2000, lambda: _tick(deadline_ticks - 1))
        except tk.TclError:
            self._run_now_poll_after = None

    def _cancel_run_now_poll(self, _evt=None):
        after_id = self._run_now_poll_after
        self._run_now_poll_after = None
        if after_id:
            try:
                self.after_cancel(after_id)
            except tk.TclError:
                pass

    def _on_reset_cache(self):
        """Clear processed_ids.json so every UNSEEN message in every account
        will be re-evaluated on the next filter run. Useful when the user
        first turns off dry-run — backlog of spam that was classified (and
        cached) during dry-run runs but never moved gets re-evaluated and
        actioned. Does NOT touch signals, whitelist, blacklist, or token
        usage — only the 'have we seen this message before' cache."""
        if not messagebox.askyesno(
                "Reset cache and re-scan inbox",
                "This clears the cache of messages MailWarden has already "
                "evaluated. The next filter run will re-check every UNSEEN "
                "message in every account.\n\n"
                "Use this once after you turn off dry run, so spam that was "
                "classified but not moved during dry run gets cleaned up.\n\n"
                "Note: already-READ messages in your inbox won't be "
                "re-scanned — the filter only looks at unread mail. If you "
                "want those acted on too, mark them unread in your email "
                "client first.\n\n"
                "Proceed?"):
            return
        try:
            config_io.save_json_atomic(
                paths.PROCESSED_IDS_PATH,
                {"version": "1.0", "ids": {}, "last_updated": ""})
        except Exception as e:
            messagebox.showerror("Could not reset cache", str(e))
            return
        # Immediately trigger a Run Now so the user sees results.
        self._on_run_now()
        messagebox.showinfo(
            "Cache cleared",
            "Filter cache cleared and a full run is starting now. Give it a "
            "minute — come back and check Filter Activity on this tab.")

    def _on_view_decisions(self):
        import re
        import tempfile
        if not paths.DECISIONS_LOG.exists():
            messagebox.showinfo(
                "No decisions yet",
                "MailWarden hasn't logged any classification decisions yet. "
                "Once the filter has evaluated some messages, they'll appear "
                "here.")
            return
        try:
            raw = paths.DECISIONS_LOG.read_text(encoding="utf-8", errors="replace")
            # Split on the '  ---' separator, reverse, rejoin newest-first.
            parts = re.split(r'\n[ \t]*---[ \t]*\n?', raw)
            blocks = [p.strip() for p in parts if p.strip()]
            reversed_text = ("\n  ---\n").join(blocks[::-1])
            tmp_path = Path(tempfile.gettempdir()) / "mailwarden-decisions-reversed.log"
            header = (
                "# Decisions log — most recent first.\n"
                "# Source: ~/MailWarden/memory/decisions.log\n"
                "# This is a generated, reversed copy for easy reading.\n\n"
            )
            tmp_path.write_text(header + reversed_text, encoding="utf-8")
            subprocess.Popen(
                ["open", str(tmp_path)],
                stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
            )
        except OSError as e:
            messagebox.showerror("Could not open log", str(e))

    def _on_open_log(self):
        if not paths.FILTER_LOG.exists():
            messagebox.showinfo(
                "No log yet",
                "The filter hasn't run yet, so there's no log to show. "
                "Click Run Now or wait for the next scheduled check.")
            return
        try:
            subprocess.Popen(["open", str(paths.FILTER_LOG)])
        except OSError as e:
            messagebox.showerror("Could not open log", str(e))

    def _maybe_show_update_banner(self):
        data = config_io.load_json(paths.UPDATE_CHECK_PATH,
                                    {"last_check": "", "latest_version": ""})
        last = data.get("last_check") or ""
        should_check = True
        if last:
            try:
                last_dt = datetime.fromisoformat(last)
                should_check = (datetime.now() - last_dt) > timedelta(days=14)
            except ValueError:
                should_check = True
        if should_check:
            latest = _fetch_latest_release_tag()
            if latest:
                data["latest_version"] = latest
            data["last_check"] = datetime.now().isoformat()
            config_io.save_json_atomic(paths.UPDATE_CHECK_PATH, data)

        latest = data.get("latest_version", "")
        if latest and _is_newer(latest, help_content.VERSION):
            self._update_banner_label.config(
                text=f"MailWarden {latest} is available (you have {help_content.VERSION}).")
            self._update_banner.pack(fill=tk.X, pady=(16, 0))
        else:
            self._update_banner.pack_forget()

    def _on_open_releases(self):
        webbrowser.open(f"{help_content.GITHUB_URL}/releases/latest")


# =============================================================================
# ACCOUNTS
# =============================================================================
class AccountsTab(ttk.Frame):
    def __init__(self, parent, app: Dashboard):
        super().__init__(parent)
        self.app = app
        self._scroll = _ScrollableTab(self)
        self._scroll.pack(fill=tk.BOTH, expand=True)
        self._f = ttk.Frame(self._scroll.body, padding=(16, 12))
        self._f.pack(fill=tk.BOTH, expand=True)
        self._build()
        self.refresh()

    def _build(self):
        columns = ("name", "email", "junk", "enabled")
        self._tree = ttk.Treeview(self._f, columns=columns, show="headings",
                                   height=14)
        for col, label, width in [
            ("name", "Name", 160),
            ("email", "Email", 260),
            ("junk", "Junk folder", 160),
            ("enabled", "Enabled", 80),
        ]:
            self._tree.heading(col, text=label)
            self._tree.column(col, width=width, stretch=(col == "email"))
        self._tree.pack(fill=tk.BOTH, expand=True)

        buttons = ttk.Frame(self._f)
        buttons.pack(fill=tk.X, pady=(8, 0))
        ttk.Button(buttons, text="Add Account…",
                   command=self._on_add).pack(side=tk.LEFT)
        ttk.Button(buttons, text="Edit…",
                   command=self._on_edit).pack(side=tk.LEFT, padx=(8, 0))
        ttk.Button(buttons, text="Remove",
                   command=self._on_remove).pack(side=tk.LEFT, padx=(8, 0))
        ttk.Button(buttons, text="Toggle Enabled",
                   command=self._on_toggle).pack(side=tk.LEFT, padx=(16, 0))

    def refresh(self):
        self._tree.delete(*self._tree.get_children())
        config = config_io.load_config()
        for i, a in enumerate(config.get("accounts", [])):
            self._tree.insert("", tk.END, iid=str(i), values=(
                a.get("name", ""),
                a.get("username", ""),
                a.get("junk_folder", "—"),
                "Yes" if a.get("enabled") else "No",
            ))

    def _selected_index(self) -> int | None:
        sel = self._tree.selection()
        if not sel:
            return None
        return int(sel[0])

    def _on_add(self):
        from . import setup_assistant
        dlg = setup_assistant.AccountFormDialog(self.app)
        self.app.wait_window(dlg)
        if dlg.saved_account is not None:
            config = config_io.load_config()
            config.setdefault("accounts", []).append(dlg.saved_account)
            config_io.save_config(config)
            self.app.refresh_all()

    def _on_edit(self):
        idx = self._selected_index()
        if idx is None:
            return
        from . import setup_assistant
        config = config_io.load_config()
        existing = config["accounts"][idx]
        dlg = setup_assistant.AccountFormDialog(self.app, existing=existing)
        self.app.wait_window(dlg)
        if dlg.saved_account is not None:
            config["accounts"][idx] = dlg.saved_account
            config_io.save_config(config)
            self.app.refresh_all()

    def _on_remove(self):
        idx = self._selected_index()
        if idx is None:
            return
        config = config_io.load_config()
        name = config["accounts"][idx]["name"]
        if not messagebox.askyesno("Remove account?",
                                    f"Remove account '{name}' from MailWarden?"):
            return
        del config["accounts"][idx]
        config_io.save_config(config)
        self.app.refresh_all()

    def _on_toggle(self):
        idx = self._selected_index()
        if idx is None:
            return
        config = config_io.load_config()
        config["accounts"][idx]["enabled"] = not config["accounts"][idx].get("enabled", True)
        config_io.save_config(config)
        self.app.refresh_all()


# =============================================================================
# WHITELIST / BLACKLIST
# =============================================================================
class ListsTab(ttk.Frame):
    def __init__(self, parent, app: Dashboard):
        super().__init__(parent)
        self.app = app
        self._scroll = _ScrollableTab(self)
        self._scroll.pack(fill=tk.BOTH, expand=True)
        self._f = ttk.Frame(self._scroll.body, padding=(16, 12))
        self._f.pack(fill=tk.BOTH, expand=True)
        self._build()
        self.refresh()

    def _build(self):
        self._conflict_label = ttk.Label(self._f, text="", foreground="#b42318")
        self._conflict_label.pack(fill=tk.X, pady=(0, 6))

        panes = ttk.PanedWindow(self._f, orient=tk.HORIZONTAL)
        panes.pack(fill=tk.BOTH, expand=True)

        # Whitelist pane
        wl_frame = ttk.LabelFrame(panes, text="Whitelist (trusted senders)",
                                   padding=(8, 6))
        panes.add(wl_frame, weight=1)
        self._wl_tree = ttk.Treeview(wl_frame, columns=("kind", "value"),
                                       show="headings")
        self._wl_tree.heading("kind", text="Type")
        self._wl_tree.heading("value", text="Value")
        self._wl_tree.column("kind", width=80, stretch=False)
        self._wl_tree.pack(fill=tk.BOTH, expand=True)
        wl_btns = ttk.Frame(wl_frame)
        wl_btns.pack(fill=tk.X, pady=(6, 0))
        ttk.Button(wl_btns, text="Add address",
                   command=lambda: self._add_entry("whitelist", "address")).pack(side=tk.LEFT)
        ttk.Button(wl_btns, text="Add domain",
                   command=lambda: self._add_entry("whitelist", "domain")).pack(side=tk.LEFT, padx=(6, 0))
        ttk.Button(wl_btns, text="Remove",
                   command=lambda: self._remove_entry("whitelist")).pack(side=tk.LEFT, padx=(12, 0))

        # Blacklist pane
        bl_frame = ttk.LabelFrame(panes, text="Blacklist (blocked senders)",
                                   padding=(8, 6))
        panes.add(bl_frame, weight=1)
        self._bl_tree = ttk.Treeview(bl_frame, columns=("kind", "value"),
                                       show="headings")
        self._bl_tree.heading("kind", text="Type")
        self._bl_tree.heading("value", text="Value")
        self._bl_tree.column("kind", width=100, stretch=False)
        self._bl_tree.pack(fill=tk.BOTH, expand=True)
        bl_btns = ttk.Frame(bl_frame)
        bl_btns.pack(fill=tk.X, pady=(6, 0))
        ttk.Button(bl_btns, text="Add address", style="Compact.TButton",
                   command=lambda: self._add_entry("blacklist", "address")).pack(side=tk.LEFT)
        ttk.Button(bl_btns, text="Add domain", style="Compact.TButton",
                   command=lambda: self._add_entry("blacklist", "domain")).pack(side=tk.LEFT, padx=(4, 0))
        ttk.Button(bl_btns, text="Add display name", style="Compact.TButton",
                   command=lambda: self._add_entry("blacklist", "display_name")).pack(side=tk.LEFT, padx=(4, 0))
        ttk.Button(bl_btns, text="Add subject keyword", style="Compact.TButton",
                   command=lambda: self._add_entry("blacklist", "subject_keyword")).pack(side=tk.LEFT, padx=(4, 0))
        ttk.Button(bl_btns, text="Remove", style="Compact.TButton",
                   command=lambda: self._remove_entry("blacklist")).pack(side=tk.LEFT, padx=(6, 0))
        ttk.Label(bl_frame, style="Muted.TLabel",
                  text="Subject keywords block any incoming email whose Subject "
                       "contains the text (case-insensitive), with no AI call.").pack(
                           anchor=tk.W, pady=(4, 0))

        # Import / export row
        io_row = ttk.Frame(self._f)
        io_row.pack(fill=tk.X, pady=(10, 0))
        ttk.Button(io_row, text="Import CSV…",
                   command=self._on_import_csv).pack(side=tk.LEFT)
        ttk.Button(io_row, text="Import XLSX…",
                   command=self._on_import_xlsx).pack(side=tk.LEFT, padx=(6, 0))
        ttk.Button(io_row, text="Export Whitelist…",
                   command=lambda: self._on_export("whitelist")).pack(side=tk.LEFT, padx=(12, 0))
        ttk.Button(io_row, text="Export Blacklist…",
                   command=lambda: self._on_export("blacklist")).pack(side=tk.LEFT, padx=(6, 0))

    def refresh(self):
        self._wl_tree.delete(*self._wl_tree.get_children())
        self._bl_tree.delete(*self._bl_tree.get_children())

        wl = config_io.load_whitelist()
        for a in wl.get("addresses", []):
            self._wl_tree.insert("", tk.END, values=("address", a))
        for d in wl.get("domains", []):
            self._wl_tree.insert("", tk.END, values=("domain", d))

        bl = config_io.load_blacklist()
        for a in bl.get("addresses", []):
            self._bl_tree.insert("", tk.END, values=("address", a))
        for d in bl.get("domains", []):
            self._bl_tree.insert("", tk.END, values=("domain", d))
        for n in bl.get("display_names", []):
            self._bl_tree.insert("", tk.END, values=("display_name", n))
        for k in bl.get("subject_keywords", []):
            self._bl_tree.insert("", tk.END, values=("subject_keyword", k))

        # Conflict detection
        addr_wl = {a.lower() for a in wl.get("addresses", [])}
        addr_bl = {a.lower() for a in bl.get("addresses", [])}
        conflicts = sorted(addr_wl & addr_bl)
        if conflicts:
            self._conflict_label.config(
                text=f"Conflict: {len(conflicts)} address(es) appear on BOTH lists — "
                     f"whitelist wins. First: {conflicts[0]}")
        else:
            self._conflict_label.config(text="")

    def _add_entry(self, which: str, kind: str):
        import re
        from tkinter import simpledialog

        MEGA_PROVIDERS = [
            "gmail.com", "yahoo.com", "outlook.com",
            "hotmail.com", "icloud.com", "aol.com",
        ]

        prompt = {"address": "Email address:",
                  "domain": "Domain to block — e.g. politicalmailers.com:",
                  "display_name": "Display name (case-insensitive):",
                  "subject_keyword": "Subject keyword to block (case-insensitive "
                                     "substring) — e.g. 'timeshare':"}[kind] \
                 if which == "blacklist" else \
                 {"address": "Email address:",
                  "domain": "Domain (without @):",
                  "display_name": "Display name (case-insensitive):"}[kind]
        val = simpledialog.askstring(f"Add to {which}", prompt, parent=self.app)
        if not val:
            return

        if kind == "domain":
            # --- Normalize ---
            val = val.strip()
            # Strip URL scheme and path
            for scheme in ("https://", "http://"):
                if val.lower().startswith(scheme):
                    val = val[len(scheme):].split("/")[0]
            # If full address pasted, take part after last @
            if "@" in val:
                val = val.rsplit("@", 1)[1]
            # Strip leading @ (defensive) and trailing dot, then lowercase
            val = val.lstrip("@").rstrip(".").lower()

            # --- Validate ---
            if "." not in val:
                messagebox.showerror("Invalid domain",
                                     f"'{val}' doesn't look like a valid domain — "
                                     "it must contain at least one dot.",
                                     parent=self.app)
                return
            if not re.fullmatch(r"[a-z0-9.\-]+", val):
                messagebox.showerror("Invalid domain",
                                     f"'{val}' contains invalid characters. "
                                     "Domains may only contain letters, numbers, dots, and hyphens.",
                                     parent=self.app)
                return

            if which == "blacklist":
                # --- Self-lockout guard ---
                cfg = config_io.load_config()
                own_domains = set()
                for acct in cfg.get("accounts", []):
                    uname = acct.get("username", "")
                    if "@" in uname:
                        own_domains.add(uname.rsplit("@", 1)[1].lower())
                if val in own_domains:
                    messagebox.showerror(
                        "Can't blacklist your own domain",
                        f"'{val}' is one of your own email domains. Adding it to the "
                        "blacklist would block your own incoming mail. "
                        "Choose a specific address instead.",
                        parent=self.app)
                    return

                # --- Whitelist conflict check ---
                wl = config_io.load_whitelist()
                if val in {d.lower() for d in wl.get("domains", [])}:
                    messagebox.showerror(
                        "Whitelist conflict",
                        f"'{val}' is on your whitelist. Remove it there first.",
                        parent=self.app)
                    return

                # --- Mega-provider warning ---
                if val in MEGA_PROVIDERS:
                    proceed = messagebox.askyesno(
                        "Block entire provider?",
                        f"Are you sure you want to block ALL mail from {val}? "
                        "This will affect every sender at that provider, including "
                        "legitimate ones. Click Yes to block anyway or No to cancel.",
                        parent=self.app)
                    if not proceed:
                        return

                # --- Deduplication ---
                data = config_io.load_blacklist()
                if val in {d.lower() for d in data.get("domains", [])}:
                    messagebox.showinfo("Already blocked",
                                        f"'{val}' is already on the blacklist.",
                                        parent=self.app)
                    return
                data.setdefault("domains", []).append(val)
                config_io.save_blacklist(data)
            else:
                # Whitelist domain
                data = config_io.load_whitelist()
                if val in {d.lower() for d in data.get("domains", [])}:
                    messagebox.showinfo("Already whitelisted",
                                        f"'{val}' is already on the whitelist.",
                                        parent=self.app)
                    return
                data.setdefault("domains", []).append(val)
                config_io.save_whitelist(data)
        elif which == "whitelist":
            val = val.strip()
            data = config_io.load_whitelist()
            key = "addresses" if kind == "address" else "domains"
            if val.lower() not in {x.lower() for x in data[key]}:
                data[key].append(val)
                config_io.save_whitelist(data)
        else:
            val = val.strip()
            data = config_io.load_blacklist()
            key = {"address": "addresses",
                   "display_name": "display_names",
                   "subject_keyword": "subject_keywords"}[kind]
            if val.lower() not in {x.lower() for x in data.setdefault(key, [])}:
                data[key].append(val)
                config_io.save_blacklist(data)
        self.refresh()

    def _remove_entry(self, which: str):
        tree = self._wl_tree if which == "whitelist" else self._bl_tree
        sel = tree.selection()
        if not sel:
            return
        kind, value = tree.item(sel[0], "values")
        if which == "whitelist":
            data = config_io.load_whitelist()
            key = "addresses" if kind == "address" else "domains"
            data[key] = [x for x in data[key] if x.lower() != value.lower()]
            config_io.save_whitelist(data)
        else:
            data = config_io.load_blacklist()
            if kind == "address":
                key = "addresses"
            elif kind == "domain":
                key = "domains"
            elif kind == "subject_keyword":
                key = "subject_keywords"
            else:
                key = "display_names"
            data[key] = [x for x in data.get(key, []) if x.lower() != value.lower()]
            config_io.save_blacklist(data)
        self.refresh()

    def _on_import_csv(self):
        _init_dir = str(paths.MAILWARDEN_ROOT) if Path(paths.MAILWARDEN_ROOT).is_dir() else str(Path.home())
        path = filedialog.askopenfilename(
            parent=self.app, title="Import CSV",
            initialdir=_init_dir,
            filetypes=[("CSV", "*.csv"), ("All files", "*.*")])
        if not path:
            return
        added, skipped = _import_tabular_csv(Path(path))
        messagebox.showinfo("Import complete",
                             f"Imported {added} new; skipped {skipped} duplicates.")
        self.refresh()

    def _on_import_xlsx(self):
        _init_dir = str(paths.MAILWARDEN_ROOT) if Path(paths.MAILWARDEN_ROOT).is_dir() else str(Path.home())
        path = filedialog.askopenfilename(
            parent=self.app, title="Import XLSX",
            initialdir=_init_dir,
            filetypes=[("Excel", "*.xlsx"), ("All files", "*.*")])
        if not path:
            return
        try:
            added, skipped = _import_tabular_xlsx(Path(path))
        except Exception as e:
            messagebox.showerror("Import failed", f"Could not read file: {e}")
            return
        messagebox.showinfo("Import complete",
                             f"Imported {added} new; skipped {skipped} duplicates.")
        self.refresh()

    def _on_export(self, which: str):
        _init_dir = str(paths.MAILWARDEN_ROOT) if Path(paths.MAILWARDEN_ROOT).is_dir() else str(Path.home())
        path = filedialog.asksaveasfilename(
            parent=self.app, title=f"Export {which}",
            initialdir=_init_dir,
            defaultextension=".csv", filetypes=[("CSV", "*.csv")])
        if not path:
            return
        data = config_io.load_whitelist() if which == "whitelist" else config_io.load_blacklist()
        with open(path, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["kind", "value"])
            if which == "whitelist":
                for a in data.get("addresses", []):
                    writer.writerow(["address", a])
                for d in data.get("domains", []):
                    writer.writerow(["domain", d])
            else:
                for a in data.get("addresses", []):
                    writer.writerow(["address", a])
                for d in data.get("domains", []):
                    writer.writerow(["domain", d])
                for n in data.get("display_names", []):
                    writer.writerow(["display_name", n])
                for k in data.get("subject_keywords", []):
                    writer.writerow(["subject_keyword", k])
        messagebox.showinfo("Export complete", f"Wrote {path}.")


# =============================================================================
# SIGNAL HISTORY
# =============================================================================
class _ScrollSection(ttk.LabelFrame):
    """A titled section with its own fixed-height scrollable body.

    Built because a single outer scroll across the whole Signal History
    tab became unusable once more than a few refinements accumulated —
    the active and pending sections would push the standard rules way
    off screen. Each section now scrolls independently so the four
    sections remain visible at a glance regardless of content volume.
    """

    def __init__(self, parent, text: str, height: int = 180):
        super().__init__(parent, text=text, padding=(6, 4))
        self._canvas = tk.Canvas(self, highlightthickness=0, borderwidth=0,
                                  height=height,
                                  background="#fafafa")
        self._scrollbar = ttk.Scrollbar(self, orient=tk.VERTICAL,
                                          command=self._canvas.yview)
        self._canvas.configure(yscrollcommand=self._scrollbar.set)
        self._canvas.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self._scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.body = ttk.Frame(self._canvas)
        self._window = self._canvas.create_window(
            (0, 0), window=self.body, anchor="nw")
        self.body.bind(
            "<Configure>",
            lambda _e: self._canvas.configure(
                scrollregion=self._canvas.bbox("all")))
        self._canvas.bind(
            "<Configure>",
            lambda e: self._canvas.itemconfig(self._window, width=e.width))

    def clear(self):
        for w in self.body.winfo_children():
            w.destroy()

    def _on_wheel(self, event):
        # macOS trackpad: event.delta is a small signed int per tick.
        # One line per tick keeps scrolling smooth — 3 lines per tick
        # (the v1.5.6 value) caused the "jumps too much" complaint.
        if event.delta == 0:
            return
        self._canvas.yview_scroll(-1 if event.delta > 0 else 1, "units")

    def bind_wheel_recursive(self):
        """Bind the section's mousewheel handler on every descendant so
        scrolling works anywhere the pointer lands inside this section.
        MUST be called after every content refresh since the old bindings
        are destroyed along with the old child widgets."""
        def _recurse(w):
            w.bind("<MouseWheel>", self._on_wheel, add="+")
            for c in w.winfo_children():
                _recurse(c)
        _recurse(self._canvas)
        _recurse(self.body)


class _ScrollableTab(ttk.Frame):
    """Standard vertical-scrolling tab body.

    Use as the outermost container for any tab that overflows the window
    height.  Add all tab widgets to `.body` exactly as you would to a
    plain ttk.Frame; the canvas and scrollbar are managed here.

    Critical constraints preserved from v1.5.7 calibration:
      - Scrollbar is always visible (never hidden-on-inactive).
      - Mouse-wheel multiplier is 1 tick per scroll event — Matt explicitly
        complained that multiplier=3 jumped too far.
      - Inner frame width tracks canvas width so horizontal content is
        never clipped on resize.
    """

    def __init__(self, parent):
        super().__init__(parent)
        try:
            bg = parent.cget("background") or "SystemWindowBackgroundColor"
        except Exception:
            bg = "SystemWindowBackgroundColor"
        self._canvas = tk.Canvas(self, highlightthickness=0, background=bg)
        vsb = ttk.Scrollbar(self, orient="vertical", command=self._canvas.yview)
        self._canvas.configure(yscrollcommand=vsb.set)

        self.body = ttk.Frame(self._canvas)
        self._window = self._canvas.create_window((0, 0), window=self.body,
                                                   anchor="nw")

        self.body.bind("<Configure>",
                       lambda e: self._canvas.configure(
                           scrollregion=self._canvas.bbox("all")))
        self._canvas.bind("<Configure>", self._on_canvas_configure)

        self._canvas.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        # Bind wheel only while the pointer is inside — avoids stealing
        # scroll from other widgets (e.g. Text widgets, Treeviews).
        self.body.bind("<Enter>",
                       lambda e: self._canvas.bind_all(
                           "<MouseWheel>", self._on_mousewheel))
        self.body.bind("<Leave>",
                       lambda e: self._canvas.unbind_all("<MouseWheel>"))

    def _on_canvas_configure(self, event):
        self._canvas.itemconfigure(self._window, width=event.width)

    def _on_mousewheel(self, event):
        # multiplier=1: v1.5.7 calibration — do not change.
        self._canvas.yview_scroll(-1 * (event.delta // 120 or 1), "units")


class SignalsTab(ttk.Frame):
    """Four-section view — each section independently scrollable:
      1. Active AI refinements
      2. Pending proposals (with Approve/Reject/Withdraw from Dashboard)
      3. Rejected / expired / withdrawn (read-only log)
      4. Standard filtering rules (hard + soft signals, not AI)
    Plus a "Forwarded spam examples" box for managing raw .eml files.
    """

    def __init__(self, parent, app: Dashboard):
        super().__init__(parent)
        self.app = app
        self._scroll = _ScrollableTab(self)
        self._scroll.pack(fill=tk.BOTH, expand=True)
        self._f = ttk.Frame(self._scroll.body, padding=(12, 8))
        self._f.pack(fill=tk.BOTH, expand=True)
        self._build_sections()
        self.refresh()

    # ---- Sections ----

    def _build_sections(self):
        ttk.Label(
            self._f, style="Muted.TLabel", wraplength=760,
            text=("MailWarden proposes refinements by email — one per new "
                  "spam pattern it learns. Reply YES/NO to those emails to "
                  "approve or reject. This tab mirrors that state so you "
                  "can see what's active, waiting on you, or rejected — "
                  "and act from here too if you prefer. Each section below "
                  "scrolls independently.")
        ).pack(anchor=tk.W, pady=(0, 8))

        # Each section is a ScrollSection (fixed height + own scrollbar).
        self._active_box = _ScrollSection(
            self._f, text="Active AI refinements (in effect)", height=200)
        self._active_box.pack(fill=tk.X, pady=(0, 8))

        self._pending_box = _ScrollSection(
            self._f, text="Pending proposals (awaiting your reply)", height=180)
        self._pending_box.pack(fill=tk.X, pady=(0, 8))

        self._history_box = _ScrollSection(
            self._f, text="Rejected / expired / withdrawn", height=140)
        self._history_box.pack(fill=tk.X, pady=(0, 8))

        self._examples_box = _ScrollSection(
            self._f, text="Forwarded spam examples (raw .eml files)", height=140)
        self._examples_box.pack(fill=tk.X, pady=(0, 8))

        self._standard_box = _ScrollSection(
            self._f, text="Standard filtering rules (hard + soft signals)",
            height=160)
        self._standard_box.pack(fill=tk.X, pady=(0, 8))

        footer = ttk.Frame(self._f)
        footer.pack(fill=tk.X, pady=(6, 2))
        ttk.Button(footer, text="Reset signals to shipped defaults…",
                   command=self._on_reset).pack(side=tk.LEFT)

    # ---- refresh & render ----

    def refresh(self):
        self._render_active()
        self._render_pending()
        self._render_history()
        self._render_examples()
        self._render_standard()
        # Re-bind wheel handlers on every section after rendering. The old
        # bindings were destroyed along with the old child widgets, so
        # scrolling would go dead after the first refresh without this.
        for section in (self._active_box, self._pending_box,
                         self._history_box, self._examples_box,
                         self._standard_box):
            section.bind_wheel_recursive()

    def _make_card(self, parent) -> ttk.Frame:
        card = ttk.Frame(parent, padding=(10, 8),
                          relief="groove", borderwidth=1)
        card.pack(fill=tk.X, pady=4)
        return card

    def _card_label(self, card, text, style=None, wraplength=720, **grid):
        kw = {"wraplength": wraplength}
        if style:
            kw["style"] = style
        lbl = ttk.Label(card, text=text, **kw)
        lbl.pack(anchor=tk.W, **grid)
        return lbl

    # ---- Active refinements ----

    def _render_active(self):
        self._active_box.clear()
        body = self._active_box.body
        refinements = sorted(
            config_io.list_active_refinements(),
            key=lambda r: r.get("last_reinforced", r.get("first_learned", "")),
            reverse=True)
        if not refinements:
            ttk.Label(
                body, style="Muted.TLabel", wraplength=720,
                text=("No active AI refinements yet. Forward a spam email "
                      "with subject \"Fwd: SPAM Example\" — MailWarden will "
                      "propose a refinement by email for your approval.")
            ).pack(anchor=tk.W)
            return
        for r in refinements:
            self._render_active_card(body, r)

    def _render_active_card(self, body, r: dict):
        card = self._make_card(body)
        headline = r.get("headline") or "(no headline)"
        self._card_label(card, headline, style="Subheading.TLabel")
        meta = (f"{r.get('kind', 'new_pattern')}  ·  "
                f"confidence {r.get('confidence', 'medium')}  ·  "
                f"matches so far: {r.get('match_count', 1)}  ·  "
                f"first learned {r.get('first_learned', '')[:16]}  ·  "
                f"last reinforced {r.get('last_reinforced', '')[:16]}  ·  "
                f"ID {r.get('id', '')}")
        self._card_label(card, meta, style="Muted.TLabel")
        if r.get("rationale"):
            self._card_label(card, f"Why: {r['rationale']}", pady=(4, 0))
        if r.get("what_this_doesnt_cover"):
            self._card_label(
                card, f"What this does NOT cover: {r['what_this_doesnt_cover']}",
                style="Muted.TLabel")
        evidence = r.get("evidence") or []
        if evidence:
            shown = ", ".join(evidence[:5])
            more = (f"  (+{len(evidence) - 5} more)"
                    if len(evidence) > 5 else "")
            self._card_label(card, f"Evidence: {shown}{more}",
                              style="Muted.TLabel")
        btns = ttk.Frame(card)
        btns.pack(anchor=tk.W, pady=(6, 0))
        rid = r.get("id", "")
        ttk.Button(btns, text="Delete",
                   command=lambda i=rid: self._on_delete_active(i)).pack(
                       side=tk.LEFT)

    def _on_delete_active(self, refinement_id: str):
        if not refinement_id:
            return
        if config_io.delete_active_refinement(refinement_id, source="dashboard"):
            self.refresh()

    # ---- Pending proposals ----

    def _render_pending(self):
        self._pending_box.clear()
        body = self._pending_box.body
        data = config_io.load_pending_signals()
        pending = [c for c in data.get("conversations", [])
                   if c.get("status") == "awaiting_reply"]
        if not pending:
            ttk.Label(
                body, style="Muted.TLabel",
                text="No proposals waiting on you.").pack(anchor=tk.W)
            return
        # Newest first
        pending.sort(key=lambda c: c.get("created", ""), reverse=True)
        for conv in pending:
            self._render_pending_card(body, conv)

    def _render_pending_card(self, body, conv: dict):
        card = self._make_card(body)
        kind = conv.get("kind", "false_positive")
        refinement = conv.get("proposed_refinement") or {}
        if kind == "spam_example_proposal":
            headline = (refinement.get("headline")
                        or f"Proposed refinement {refinement.get('id', '')}")
        else:
            headline = (f"False-positive narrowing for "
                        f"{conv.get('original_from', 'unknown sender')}")
        self._card_label(card, headline, style="Subheading.TLabel")

        # Show hard/soft type badge for spam_example_proposals
        if kind == "spam_example_proposal":
            signal_type = refinement.get("signal_type", "soft")
            hard_rule = refinement.get("hard_rule") or {}
            if signal_type == "hard" and hard_rule:
                hr_type = hard_rule.get("type", "")
                hr_value = hard_rule.get("value", "")
                if hr_type == "subject_keyword":
                    type_text = (f"Hard rule — no AI cost  "
                                 f"(blocks subject containing \"{hr_value}\")")
                else:
                    type_text = (f"Hard rule — no AI cost  "
                                 f"(blocks sender domain {hr_value})")
                ttk.Label(card, text=type_text,
                          foreground="#1a5fa8",
                          wraplength=720).pack(anchor=tk.W)
            else:
                ttk.Label(card, text="AI refinement (soft)",
                          foreground="#555555",
                          wraplength=720).pack(anchor=tk.W)

        meta_parts = [f"SFID {conv.get('id', '')}",
                       f"kind: {kind}",
                       f"created {conv.get('created', '')[:16]}",
                       f"expires {conv.get('expires', '')[:10]}"]
        if refinement.get("confidence"):
            meta_parts.append(f"confidence {refinement['confidence']}")
        self._card_label(card, "  ·  ".join(meta_parts), style="Muted.TLabel")
        if refinement.get("rationale"):
            self._card_label(card, f"Why: {refinement['rationale']}",
                              pady=(4, 0))
        if refinement.get("what_this_doesnt_cover"):
            self._card_label(
                card, f"What this does NOT cover: "
                      f"{refinement['what_this_doesnt_cover']}",
                style="Muted.TLabel")
        if refinement.get("evidence"):
            self._card_label(card,
                              f"Evidence: {', '.join(refinement['evidence'][:5])}",
                              style="Muted.TLabel")
        elif conv.get("original_subject"):
            self._card_label(
                card, f"Evidence subject: {conv['original_subject']}",
                style="Muted.TLabel")

        btns = ttk.Frame(card)
        btns.pack(anchor=tk.W, pady=(8, 0))
        sfid = conv.get("id", "")
        if kind == "spam_example_proposal":
            signal_type = refinement.get("signal_type", "soft")
            ttk.Button(btns, text="Approve", style="Primary.TButton",
                       command=lambda s=sfid: self._on_approve_pending(s)).pack(
                           side=tk.LEFT)
            # Downgrade button: for hard proposals, offer "Apply as soft instead"
            if signal_type == "hard":
                ttk.Button(btns, text="Apply as soft instead",
                           command=lambda s=sfid: self._on_approve_as_soft(s)).pack(
                               side=tk.LEFT, padx=(6, 0))
            ttk.Button(btns, text="Reject",
                       command=lambda s=sfid: self._on_reject_pending(s)).pack(
                           side=tk.LEFT, padx=(6, 0))
        else:
            ttk.Label(btns, style="Muted.TLabel",
                      text="(False-positive narrowings: reply to the email "
                            "to approve.)").pack(side=tk.LEFT)
        ttk.Button(btns, text="Withdraw",
                   command=lambda s=sfid: self._on_withdraw_pending(s)).pack(
                       side=tk.LEFT, padx=(6, 0))

    @staticmethod
    def _apply_hard_rule_to_blacklist(hard_rule: dict) -> str | None:
        """Write a hard_rule into blacklist.json atomically.

        Returns a human-readable description of what was added, or None if
        the rule was already present (de-duped) or if the shape was invalid.
        Raises on unexpected IO errors so the caller can surface them.
        """
        hr_type = (hard_rule.get("type") or "").strip().lower()
        value = (hard_rule.get("value") or "").strip()
        if not value or hr_type not in ("subject_keyword", "sender_domain"):
            return None

        bl = config_io.load_blacklist()

        if hr_type == "subject_keyword":
            existing = {k.lower() for k in bl.get("subject_keywords", [])}
            if value.lower() in existing:
                return None  # already present
            bl.setdefault("subject_keywords", []).append(value)
            config_io.save_blacklist(bl)
            return f"subject keyword \"{value}\""

        # sender_domain: normalise (strip @) then de-dupe
        domain = value.lstrip("@").lower()
        existing = {d.lower().lstrip("@") for d in bl.get("domains", [])}
        if domain in existing:
            return None  # already present
        bl.setdefault("domains", []).append(domain)
        config_io.save_blacklist(bl)
        return f"sender domain {domain}"

    def _on_approve_pending(self, sfid: str):
        """Approve a pending proposal.

        For HARD proposals: write the hard_rule to blacklist.json, then mark
        the conversation approved (same as soft, so history is consistent).
        For SOFT proposals: existing behaviour — activate the ai_refinement.
        """
        pending = config_io.load_pending_signals()
        conv = next((c for c in pending.get("conversations", [])
                     if c.get("id") == sfid), None)
        if conv is None:
            messagebox.showerror(
                "Could not apply",
                f"SFID {sfid} not found or not approvable from the Dashboard "
                f"(false-positive narrowings must be approved by email reply).")
            self.refresh()
            return

        refinement = conv.get("proposed_refinement") or {}
        signal_type = refinement.get("signal_type", "soft")
        hard_rule = refinement.get("hard_rule") if signal_type == "hard" else None

        if hard_rule:
            # Hard path: apply to blacklist, then mark approved in pending
            try:
                added = self._apply_hard_rule_to_blacklist(hard_rule)
            except Exception as e:
                messagebox.showerror("Could not apply hard rule",
                                     f"Failed to update blacklist: {e}")
                self.refresh()
                return

            # Mark the conversation resolved in pending_signals
            conv["status"] = "approved"
            conv["resolution"] = "approved"
            conv.setdefault("conversation_history", []).append({
                "role": "system",
                "timestamp": config_io.now_iso(),
                "content": "Approved (hard rule) via dashboard",
            })
            config_io.save_pending_signals(pending)
            config_io.append_refinement_log({
                "ts": config_io.now_iso(),
                "event": "applied",
                "id": refinement.get("id", ""),
                "sfid": sfid,
                "headline": refinement.get("headline", ""),
                "signal_type": "hard",
                "hard_rule": hard_rule,
                "source": "dashboard",
            })

            if added:
                messagebox.showinfo(
                    "Hard rule applied",
                    f"Blacklist updated — added {added}.\n"
                    f"Mail matching this rule will be blocked instantly, "
                    f"with no AI call.")
            else:
                messagebox.showinfo(
                    "Already blocked",
                    f"The rule was already in the blacklist. No change needed.")
        else:
            # Soft path: existing behaviour
            applied = config_io.apply_refinement_from_pending(
                sfid, source="dashboard")
            if applied:
                messagebox.showinfo(
                    "Applied",
                    f"Refinement {applied.get('id', '')} is now active.")
            else:
                messagebox.showerror(
                    "Could not apply",
                    f"SFID {sfid} not found or not approvable from the Dashboard "
                    f"(false-positive narrowings must be approved by email reply).")
        self.refresh()

    def _on_approve_as_soft(self, sfid: str):
        """Downgrade a hard proposal and apply it as a soft AI refinement."""
        pending = config_io.load_pending_signals()
        conv = next((c for c in pending.get("conversations", [])
                     if c.get("id") == sfid), None)
        if conv is None:
            messagebox.showerror("Not found", f"SFID {sfid} not found.")
            self.refresh()
            return

        refinement = conv.get("proposed_refinement") or {}
        if refinement.get("signal_type") != "hard":
            # Nothing to downgrade — just do a normal approve
            self._on_approve_pending(sfid)
            return

        # Strip the hard_rule fields so apply_refinement_from_pending treats
        # this as a plain soft refinement
        refinement["signal_type"] = "soft"
        refinement.pop("hard_rule", None)
        config_io.save_pending_signals(pending)

        applied = config_io.apply_refinement_from_pending(sfid, source="dashboard")
        if applied:
            messagebox.showinfo(
                "Applied as soft",
                f"Refinement {applied.get('id', '')} is now active as an "
                f"AI refinement (soft).")
        else:
            messagebox.showerror(
                "Could not apply",
                f"SFID {sfid} could not be applied.")
        self.refresh()

    def _on_reject_pending(self, sfid: str):
        if config_io.reject_pending(sfid, source="dashboard"):
            self.refresh()
        else:
            messagebox.showerror("Could not reject",
                                  f"SFID {sfid} not found.")

    def _on_withdraw_pending(self, sfid: str):
        if not messagebox.askyesno(
            "Withdraw proposal",
            f"Remove pending proposal {sfid}? You can always forward the "
            "example again to re-propose."):
            return
        if config_io.withdraw_pending(sfid, source="dashboard"):
            self.refresh()

    # ---- History ----

    def _render_history(self):
        self._history_box.clear()
        body = self._history_box.body
        log = config_io.load_refinement_log(limit=200)
        shown_events = {"rejected", "expired", "withdrawn", "deleted"}
        entries = [e for e in log if e.get("event") in shown_events]
        if not entries:
            ttk.Label(body, style="Muted.TLabel",
                      text="No rejected, expired, or deleted refinements yet."
                      ).pack(anchor=tk.W)
            return
        for e in entries[:50]:
            card = ttk.Frame(body, padding=(8, 4))
            card.pack(fill=tk.X, pady=2)
            first = (f"{e.get('ts', '')[:16]}  ·  {e.get('event', '').upper()}"
                     f"  ·  {e.get('headline', '') or e.get('id', '') or e.get('sfid', '')}")
            ttk.Label(card, text=first, style="Muted.TLabel",
                      wraplength=720).pack(anchor=tk.W)
            if e.get("source") or e.get("reason"):
                extra = []
                if e.get("source"):
                    extra.append(f"via {e['source']}")
                if e.get("reason"):
                    extra.append(e["reason"])
                ttk.Label(card, text="  ·  ".join(extra), style="Muted.TLabel",
                          wraplength=720).pack(anchor=tk.W)

    # ---- Spam examples (raw files) ----

    def _render_examples(self):
        self._examples_box.clear()
        body = self._examples_box.body
        if not paths.SPAM_EXAMPLES_DIR.exists():
            ttk.Label(body, style="Muted.TLabel",
                      text="No spam_examples folder yet.").pack(anchor=tk.W)
            return
        files = sorted(paths.SPAM_EXAMPLES_DIR.glob("*.eml"))
        if not files:
            ttk.Label(body, style="Muted.TLabel",
                      text="No .eml files yet — forward a spam with "
                            "subject \"Fwd: SPAM Example\" to populate this."
                      ).pack(anchor=tk.W)
            return
        ttk.Label(body, style="Muted.TLabel", wraplength=720,
                  text=f"{len(files)} file(s). Deleting a file makes the "
                        f"learner ignore it on future runs; already-applied "
                        f"refinements stay.").pack(anchor=tk.W)
        row_wrap = ttk.Frame(body)
        row_wrap.pack(fill=tk.X, pady=(4, 0))
        self._ex_list = tk.Listbox(row_wrap, height=min(8, max(3, len(files))),
                                     background="#ffffff",
                                     foreground="#111111",
                                     selectbackground="#c0d0ff")
        self._ex_list.pack(side=tk.LEFT, fill=tk.X, expand=True)
        for p in files:
            self._ex_list.insert(tk.END, p.name)
        btns = ttk.Frame(row_wrap)
        btns.pack(side=tk.RIGHT, padx=(8, 0))
        ttk.Button(btns, text="Open",
                   command=self._on_open_example).pack(fill=tk.X)
        ttk.Button(btns, text="Delete",
                   command=self._on_delete_example).pack(fill=tk.X,
                                                           pady=(4, 0))

    def _on_open_example(self):
        if not hasattr(self, "_ex_list"):
            return
        sel = self._ex_list.curselection()
        if not sel:
            return
        name = self._ex_list.get(sel[0])
        try:
            subprocess.Popen(["open", str(paths.SPAM_EXAMPLES_DIR / name)])
        except OSError as e:
            messagebox.showerror("Could not open", str(e))

    def _on_delete_example(self):
        if not hasattr(self, "_ex_list"):
            return
        sel = self._ex_list.curselection()
        if not sel:
            return
        name = self._ex_list.get(sel[0])
        if not messagebox.askyesno(
            "Delete example",
            f"Delete {name}? The learner will no longer consider this file."):
            return
        (paths.SPAM_EXAMPLES_DIR / name).unlink(missing_ok=True)
        self.refresh()

    # ---- Standard rules ----

    def _render_standard(self):
        self._standard_box.clear()
        body = self._standard_box.body
        sig_data = config_io.load_signals().get("signals", {})
        ttk.Label(
            body, style="Muted.TLabel", wraplength=720,
            text=("SpamAssassin-style flags the filter computes from headers "
                  "and body. Hard signals move mail to Junk without calling "
                  "Claude. Soft signals are passed to Claude as context — "
                  "Claude weighs them alongside the message content when "
                  "deciding. These flags themselves aren't AI-learned; "
                  "they're listed here for transparency.")
        ).pack(anchor=tk.W, pady=(0, 4))
        self._render_rule_group(
            body, "Hard signals (auto-spam, no API call)",
            sig_data.get("hard_signals", []))
        self._render_rule_group(
            body, "Soft signals (weighted by Claude at classification)",
            sig_data.get("soft_signals", []))
        self._render_rule_group(
            body, "Known sending infrastructure",
            sig_data.get("known_sending_infrastructure", []))
        self._render_rule_group(
            body, "Known impersonated brands",
            sig_data.get("known_impersonated_brands", []))
        notes = sig_data.get("learner_notes", "")
        if notes:
            ttk.Label(body, style="Muted.TLabel",
                      text="Legacy learner notes (pre-refinement era):",
                      wraplength=720).pack(anchor=tk.W, pady=(6, 0))
            ttk.Label(body, text=notes, wraplength=720).pack(anchor=tk.W)

    def _render_rule_group(self, body, title: str, items: list):
        header = f"{title} — {len(items)}"
        ttk.Label(body, text=header,
                  style="Subheading.TLabel").pack(anchor=tk.W, pady=(6, 0))
        if not items:
            ttk.Label(body, text="  (none)",
                      style="Muted.TLabel").pack(anchor=tk.W)
            return
        for item in items[:25]:
            ttk.Label(body, text=f"  • {item}",
                      wraplength=720).pack(anchor=tk.W)
        if len(items) > 25:
            ttk.Label(body,
                      text=f"  … +{len(items) - 25} more",
                      style="Muted.TLabel").pack(anchor=tk.W)

    # ---- Footer actions ----

    def _on_reset(self):
        if not _triple_confirm(
            self.app,
            "Reset signals to shipped defaults",
            "This replaces your learned spam signals with the defaults bundled "
            "with MailWarden. All accumulated AI refinements, infrastructure "
            "patterns, and narrowings will be lost."):
            return
        from . import app_entrypoint
        default_src = app_entrypoint.get_bundled_defaults_dir() / "signals.json"
        if not default_src.exists():
            messagebox.showerror("Defaults missing",
                                  f"Cannot find {default_src}.")
            return
        import shutil
        shutil.copy2(default_src, paths.SIGNALS_PATH)
        messagebox.showinfo("Reset",
                             "Signals reset to shipped defaults.")
        self.refresh()


# =============================================================================
# API USAGE
# =============================================================================
class UsageTab(ttk.Frame):
    def __init__(self, parent, app: Dashboard):
        super().__init__(parent)
        self.app = app
        self._labels: dict[str, ttk.Label] = {}
        self._scroll = _ScrollableTab(self)
        self._scroll.pack(fill=tk.BOTH, expand=True)
        self._f = ttk.Frame(self._scroll.body, padding=(16, 12))
        self._f.pack(fill=tk.BOTH, expand=True)
        self._build()
        self.refresh()

    def _build(self):
        # Cost/dollar estimates were removed for the beta: a normal Anthropic
        # API key can't query billing, so the figures were unverifiable. This
        # tab now shows only factual, non-dollar activity counts. Monitor
        # actual spend in the Anthropic console.
        frm = ttk.LabelFrame(self._f, text="Classifier activity",
                              padding=(10, 8))
        frm.pack(fill=tk.X)
        rows = ("Emails classified — last 24 hours",
                "Emails classified — last 7 days",
                "Emails classified — month-to-date",
                "Emails classified — lifetime",
                "Pre-classifier skips (lifetime)")
        for i, label in enumerate(rows):
            ttk.Label(frm, text=f"{label}:").grid(row=i, column=0, sticky=tk.W, pady=2)
            v = ttk.Label(frm, text="—")
            v.grid(row=i, column=1, sticky=tk.W, padx=(16, 0), pady=2)
            self._labels[label] = v

        ttk.Label(self._f, style="Muted.TLabel", wraplength=720, text=(
            "Pre-classifier skips are emails decided by cheap local checks "
            "without calling the AI. To see exact dollar charges, open the "
            "Anthropic console.")).pack(anchor=tk.W, pady=(10, 0))

        actions = ttk.Frame(self._f)
        actions.pack(fill=tk.X, pady=(16, 0))
        ttk.Button(actions, text="Open Anthropic Console",
                   command=lambda: webbrowser.open(help_content.ANTHROPIC_BILLING_URL)).pack(side=tk.LEFT)

    def refresh(self):
        """Read token_usage.json in the schema the filter actually writes:
            lifetime_api_calls
            daily_records: list of {date, api_calls,
                                    api_calls_skipped_by_pre_classifier}
        Only non-dollar activity counts are surfaced.
        """
        usage = config_io.load_token_usage()
        daily_records = usage.get("daily_records", []) or []

        # Pre-classifier skip count: either explicit lifetime total, or the
        # sum of the per-day skip counts.
        skips = int(usage.get("pre_classifier_skips_lifetime",
                               usage.get("pre_classifier_skips", 0)))
        if not skips:
            skips = sum(int(d.get("api_calls_skipped_by_pre_classifier", 0))
                         for d in daily_records)

        # Index records by parsed date for fast lookup.
        by_date: dict = {}
        for d in daily_records:
            try:
                k = datetime.strptime(d.get("date", ""), "%Y-%m-%d").date()
            except (ValueError, TypeError):
                continue
            by_date[k] = d

        today = datetime.now().date()
        week_start = today - timedelta(days=6)
        month_prefix = today.strftime("%Y-%m")

        def _calls(predicate):
            return sum(int(rec.get("api_calls", 0))
                       for dt, rec in by_date.items() if predicate(dt))

        n24 = _calls(lambda dt: dt == today)
        n7 = _calls(lambda dt: dt >= week_start)
        nmtd = _calls(lambda dt: dt.strftime("%Y-%m") == month_prefix)
        lifetime_calls = int(usage.get("lifetime_api_calls", 0)) or _calls(
            lambda dt: True)

        self._labels["Emails classified — last 24 hours"].config(text=f"{n24:,}")
        self._labels["Emails classified — last 7 days"].config(text=f"{n7:,}")
        self._labels["Emails classified — month-to-date"].config(text=f"{nmtd:,}")
        self._labels["Emails classified — lifetime"].config(text=f"{lifetime_calls:,}")
        self._labels["Pre-classifier skips (lifetime)"].config(text=f"{skips:,}")


# =============================================================================
# SETTINGS
# =============================================================================
class SettingsTab(ttk.Frame):
    def __init__(self, parent, app: Dashboard):
        super().__init__(parent)
        self.app = app
        self._scroll = _ScrollableTab(self)
        self._scroll.pack(fill=tk.BOTH, expand=True)
        self._f = ttk.Frame(self._scroll.body, padding=(16, 12))
        self._f.pack(fill=tk.BOTH, expand=True)
        self._build()
        self.refresh()

    def _build(self):
        # API key
        api_frame = ttk.LabelFrame(self._f, text="Anthropic API key",
                                    padding=(10, 6))
        api_frame.pack(fill=tk.X, pady=(0, 8))
        self._api_var = tk.StringVar()
        entry = ttk.Entry(api_frame, textvariable=self._api_var, show="•", width=60)
        entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Button(api_frame, text="Save",
                   command=self._on_save_api).pack(side=tk.LEFT, padx=(8, 0))
        ttk.Button(api_frame, text="Validate",
                   command=self._on_validate_api).pack(side=tk.LEFT, padx=(4, 0))
        self._api_status = ttk.Label(api_frame, text="", foreground="#555")
        self._api_status.pack(side=tk.LEFT, padx=(12, 0))

        # Model + threshold + max emails
        ai_frame = ttk.LabelFrame(self._f, text="Classifier", padding=(10, 6))
        ai_frame.pack(fill=tk.X, pady=(0, 8))

        ttk.Label(ai_frame, text="Model:").grid(row=0, column=0, sticky=tk.W, padx=(0, 8), pady=4)
        self._model_var = tk.StringVar()
        self._model_box = ttk.Combobox(ai_frame, textvariable=self._model_var,
                                         values=[m[0] for m in MODEL_CHOICES],
                                         state="readonly", width=36)
        self._model_box.grid(row=0, column=1, sticky=tk.W, pady=4)
        self._model_box.bind("<<ComboboxSelected>>", self._on_model_selected)
        # Transient "✓ Saved" confirmation shown next to the dropdown after an
        # instant save. Lives in column 2 so it doesn't shift the dropdown.
        self._model_saved_label = ttk.Label(ai_frame, text="", foreground="#2a7")
        self._model_saved_label.grid(row=0, column=2, sticky=tk.W, padx=(8, 0))
        self._model_saved_after = None

        ttk.Label(ai_frame, text="Confidence threshold:").grid(row=1, column=0, sticky=tk.W, padx=(0, 8), pady=4)
        self._threshold_var = tk.DoubleVar(value=0.85)
        scale = ttk.Scale(ai_frame, from_=0.70, to=0.99, orient=tk.HORIZONTAL,
                           variable=self._threshold_var, length=240,
                           command=lambda _v: self._threshold_label.config(
                               text=f"{self._threshold_var.get():.2f}"))
        scale.grid(row=1, column=1, sticky=tk.W, pady=4)
        self._threshold_label = ttk.Label(ai_frame, text="0.85")
        self._threshold_label.grid(row=1, column=2, padx=(8, 0))

        ttk.Label(ai_frame, text="Max emails per run:").grid(row=2, column=0, sticky=tk.W, padx=(0, 8), pady=4)
        self._maxrun_var = tk.IntVar(value=100)
        ttk.Spinbox(ai_frame, from_=1, to=1000, textvariable=self._maxrun_var,
                     width=10).grid(row=2, column=1, sticky=tk.W, pady=4)

        # Menu bar + report + global pause
        misc = ttk.LabelFrame(self._f, text="Menu bar and schedule",
                               padding=(10, 6))
        misc.pack(fill=tk.X, pady=(0, 8))

        self._menubar_var = tk.BooleanVar()
        ttk.Checkbutton(misc, text="Show MailWarden in the menu bar",
                        variable=self._menubar_var,
                        command=self._on_menubar_toggle).pack(anchor=tk.W)

        self._train_prompt_var = tk.BooleanVar()
        ttk.Checkbutton(
            misc,
            text="On launch, offer to create the \"Train MailWarden\" "
                  "folder in any accounts that don't have it",
            variable=self._train_prompt_var,
            command=self._on_train_prompt_toggle
        ).pack(anchor=tk.W, pady=(2, 0))
        ttk.Button(
            misc, text="Check Train folders now",
            command=self._on_check_train_folders_now
        ).pack(anchor=tk.W, pady=(2, 0))

        time_row = ttk.Frame(misc)
        time_row.pack(fill=tk.X, pady=(6, 0))
        ttk.Label(time_row, text="Daily report time (24h HH:MM):").pack(side=tk.LEFT)
        self._report_time_var = tk.StringVar(value="08:00")
        ttk.Entry(time_row, textvariable=self._report_time_var, width=8).pack(side=tk.LEFT, padx=(8, 0))
        ttk.Button(time_row, text="Apply",
                   command=self._on_apply_report_time).pack(side=tk.LEFT, padx=(8, 0))

        # Filter check interval — how often launchd wakes spam_filter.py.
        interval_row = ttk.Frame(misc)
        interval_row.pack(fill=tk.X, pady=(6, 0))
        ttk.Label(interval_row, text="Check inbox every:").pack(side=tk.LEFT)
        self._interval_var = tk.IntVar(value=15)
        ttk.Spinbox(interval_row, from_=5, to=360, increment=5,
                    textvariable=self._interval_var, width=6).pack(
                        side=tk.LEFT, padx=(8, 0))
        ttk.Label(interval_row, text="minutes").pack(side=tk.LEFT, padx=(4, 0))
        ttk.Button(interval_row, text="Apply",
                   command=self._on_apply_filter_interval).pack(
                       side=tk.LEFT, padx=(8, 0))
        ttk.Label(misc, style="Muted.TLabel", wraplength=560,
                  text=("Shorter intervals catch spam faster but use more "
                        "Anthropic API credit. 15 minutes is the default; "
                        "5 is the minimum; 60 is plenty for low-volume "
                        "inboxes.")).pack(
                            anchor=tk.W, pady=(2, 0), padx=(2, 0))

        pause_row = ttk.Frame(misc)
        pause_row.pack(fill=tk.X, pady=(6, 0))
        self._pause_btn = ttk.Button(pause_row, text="Pause all filtering",
                                        command=self._on_pause_toggle)
        self._pause_btn.pack(side=tk.LEFT)
        # Nuclear-option rescue: fully bootout, rewrite, and re-bootstrap every
        # MailWarden launchd agent. Useful when Sonoma's launchd has ghost
        # entries from an older install that the normal self-heal can't clear.
        ttk.Button(pause_row, text="Restart all background services",
                   command=self._on_rebuild_schedulers).pack(
                       side=tk.LEFT, padx=(8, 0))

        # Actions
        contribute = ttk.LabelFrame(self._f, text="Contribute learned signals",
                                    padding=(10, 6))
        contribute.pack(fill=tk.X, pady=(0, 8))
        ttk.Label(contribute, text=(
            "Help improve MailWarden by sharing your learned spam patterns. "
            "This only sends signal descriptions, never your emails or contacts. "
            "Opt in anytime."), wraplength=620).pack(anchor=tk.W)
        ttk.Button(contribute, text="Share my learned signals",
                   command=self._on_share_signals).pack(anchor=tk.W, pady=(6, 0))

        # Danger zone — full uninstall. Placed last so it sits at the bottom
        # of the scrollable Settings tab, visually set apart as destructive.
        danger = ttk.LabelFrame(self._f, text="Uninstall", padding=(10, 6))
        danger.pack(fill=tk.X, pady=(16, 8))
        ttk.Label(danger, wraplength=620, text=(
            "Completely remove MailWarden from this Mac. This unregisters "
            "MailWarden's background services and moves the app to the Trash. "
            "You can optionally also delete your settings, history, and saved "
            "passwords.")).pack(anchor=tk.W)
        ttk.Button(danger, text="Uninstall MailWarden…", style="Danger.TButton",
                   command=self._on_uninstall).pack(anchor=tk.W, pady=(8, 0))

    def refresh(self):
        config = config_io.load_config()
        self._api_var.set(config.get("anthropic", {}).get("api_key", ""))
        default_model = "claude-haiku-4-5-20251001"
        current_model = config.get("anthropic", {}).get("model", default_model)
        for label, value in MODEL_CHOICES:
            if value == current_model:
                self._model_var.set(label)
                break
        else:
            # Unmatched/missing model → fall back to the Haiku entry (the
            # functional default), never Sonnet. Look it up by value so this
            # stays correct even if MODEL_CHOICES order changes.
            self._model_var.set(
                next(l for l, v in MODEL_CHOICES if v == default_model))
        self._threshold_var.set(config.get("anthropic", {}).get("confidence_threshold", 0.85))
        self._threshold_label.config(text=f"{self._threshold_var.get():.2f}")
        self._maxrun_var.set(config.get("filter", {}).get("max_emails_per_run", 100))
        self._menubar_var.set(config.get("ui", {}).get("menu_bar_enabled", True))
        self._train_prompt_var.set(
            config.get("ui", {}).get("prompt_missing_train_folder", True))

        hour = config.get("summary", {}).get("hour", 8)
        minute = config.get("summary", {}).get("minute", 0)
        self._report_time_var.set(f"{hour:02d}:{minute:02d}")
        self._interval_var.set(
            int(config.get("filter", {}).get("interval_minutes", 15))
        )

        paused = config.get("ui", {}).get("paused", False) or (
            config.get("accounts") and not any(a.get("enabled") for a in config["accounts"]))
        self._pause_btn.config(text=("Resume all filtering" if paused else "Pause all filtering"))

    def _on_save_api(self):
        config = config_io.load_config()
        config.setdefault("anthropic", {})["api_key"] = self._api_var.get().strip()
        # Also save model + threshold + max
        for label, value in MODEL_CHOICES:
            if label == self._model_var.get():
                config["anthropic"]["model"] = value
                break
        config["anthropic"]["confidence_threshold"] = round(float(self._threshold_var.get()), 2)
        config.setdefault("filter", {})["max_emails_per_run"] = int(self._maxrun_var.get())
        config_io.save_config(config)
        self._api_status.config(text="Saved.")

    def _on_model_selected(self, _event=None):
        """Persist the model immediately when the dropdown changes.

        Maps the selected display label back to its model VALUE via
        MODEL_CHOICES (never saves the human label), then reuses the same
        config-save mechanism as the API-row Save button so the two paths
        always agree on key and storage.
        """
        selected = self._model_var.get()
        model_value = next(
            (v for label, v in MODEL_CHOICES if label == selected), None)
        if model_value is None:
            return
        config = config_io.load_config()
        config.setdefault("anthropic", {})["model"] = model_value
        config_io.save_config(config)
        self._show_model_saved()

    def _show_model_saved(self):
        """Briefly show a non-modal '✓ Saved' next to the dropdown.

        Cancels any pending timer first so a destroyed widget is never
        touched by a stale after() callback.
        """
        if self._model_saved_after is not None:
            self.after_cancel(self._model_saved_after)
            self._model_saved_after = None
        self._model_saved_label.config(text="✓ Saved")

        def clear():
            self._model_saved_after = None
            if self._model_saved_label.winfo_exists():
                self._model_saved_label.config(text="")

        self._model_saved_after = self.after(2000, clear)

    def _on_validate_api(self):
        import threading
        self._api_status.config(text="Validating…")
        key = self._api_var.get().strip()

        def worker():
            ok, msg = validators.validate_api_key(key)

            def done():
                self._api_status.config(text=msg)

            self.after(0, done)

        threading.Thread(target=worker, daemon=True).start()

    def _on_menubar_toggle(self):
        # v1.6.0: SMAppService manages agent lifecycle. We write the user's
        # preference to config.json so the filter/report agents respect it,
        # and then register or unregister via SMAppService. No launchctl or
        # plist writing needed.
        config = config_io.load_config()
        on = bool(self._menubar_var.get())
        config.setdefault("ui", {})["menu_bar_enabled"] = on
        config_io.save_config(config)
        try:
            if on:
                ok, err = smappservice_install.register_menubar()
                if not ok and err:
                    messagebox.showerror("Menu bar", str(err))
            else:
                ok, err = smappservice_install.unregister_menubar()
                if not ok and err:
                    messagebox.showerror("Menu bar", str(err))
        except Exception as e:
            messagebox.showerror("Menu bar", str(e))

    def _on_train_prompt_toggle(self):
        config = config_io.load_config()
        config.setdefault("ui", {})["prompt_missing_train_folder"] = \
            bool(self._train_prompt_var.get())
        config_io.save_config(config)

    def _on_check_train_folders_now(self):
        # Button bypasses the prompt-on-launch setting — user explicitly asked.
        import threading
        config = config_io.load_config()
        threading.Thread(
            target=self.app._check_train_folders_bg,
            args=(config,), daemon=True).start()

    def _on_apply_report_time(self):
        try:
            hour, _, minute = self._report_time_var.get().partition(":")
            h = int(hour)
            m = int(minute) if minute else 0
        except ValueError:
            messagebox.showerror("Time format", "Use HH:MM in 24-hour format.")
            return
        config = config_io.load_config()
        config.setdefault("summary", {})["hour"] = h
        config["summary"]["minute"] = m
        config_io.save_config(config)
        # v1.6.0: SMAppService plists are static inside the .app bundle.
        # Schedule changes are written to config.json; daily_report.py reads
        # config.json at runtime for its report hour/minute. No plist rewrite
        # or agent re-registration needed.
        messagebox.showinfo("Applied",
                             f"Daily report will now run at {h:02d}:{m:02d}.")

    def _on_apply_filter_interval(self):
        try:
            minutes = int(self._interval_var.get())
        except (ValueError, tk.TclError):
            messagebox.showerror(
                "Interval",
                "Enter a whole number of minutes between 5 and 360.")
            return
        if minutes < 5 or minutes > 360:
            messagebox.showerror(
                "Interval",
                "Choose a value between 5 minutes and 360 minutes (6 hours).")
            return
        # Light confirmation for very frequent intervals. This does NOT cost
        # extra API credit (already-classified mail is cached and never re-sent
        # to the classifier), but very frequent fresh logins can trip provider
        # rate limits.
        if minutes < 10:
            proceed = messagebox.askyesno(
                "High-frequency checking",
                f"Checking every {minutes} minutes can cause your email "
                f"provider to throttle or temporarily block sign-ins — "
                f"MailWarden logs in fresh every run, and AOL and Yahoo in "
                f"particular react badly to very frequent logins.\n\n"
                f"Continue?")
            if not proceed:
                return
        config = config_io.load_config()
        config.setdefault("filter", {})["interval_minutes"] = minutes
        config_io.save_config(config)
        # v1.6.0: SMAppService plists are static and signed read-only, so the
        # cadence can't be written into the plist. Instead the plist wakes the
        # filter every 5 minutes and spam_filter.run_filter() reads this value
        # at runtime: an elapsed-time gate skips any scheduled wake that fires
        # sooner than interval_minutes after the last real run. No plist
        # rewrite or agent re-registration needed.
        messagebox.showinfo("Applied",
                             f"Filter will now run every {minutes} minutes.")

    def _on_rebuild_schedulers(self):
        """Nuclear-option rescue: unregister then re-register MailWarden's
        background agents via SMAppService. User's escape hatch for when
        the agents have stopped responding.

        v1.6.0: uses SMAppService unregister + register instead of launchctl
        bootout + bootstrap. The worker-thread + progressbar + result-
        messagebox pattern is preserved verbatim from v1.5.14.
        """
        import threading
        if not messagebox.askyesno(
            "Restart all background services",
            "This fully unregisters and re-registers MailWarden's background "
            "services (the scheduled spam filter, daily report, and menu "
            "bar agent). Use it if scheduled runs have stopped firing or "
            "the menu bar icon has disappeared. Continue?"):
            return

        # Snapshot user settings on the Tk thread (Tk vars can't be read
        # safely from a worker thread).
        try:
            config = config_io.load_config()
            install_menubar = bool(self._menubar_var.get())
        except Exception as e:
            messagebox.showerror("Rebuild failed",
                                 f"Could not read settings: "
                                 f"{type(e).__name__}: {e}")
            return

        # Modal progress dialog — same pattern as v1.5.14 to keep UX consistent.
        top = tk.Toplevel(self)
        top.title("Restarting background services")
        top.transient(self.winfo_toplevel())
        top.grab_set()
        top.resizable(False, False)
        ttk.Label(top,
                  text=("Restarting MailWarden's scheduled agents.\n"
                        "This may take a few seconds..."),
                  padding=(20, 16)).pack()
        pb = ttk.Progressbar(top, mode="indeterminate", length=320)
        pb.pack(padx=20, pady=(0, 16))
        pb.start(80)

        top.update_idletasks()
        try:
            parent = self.winfo_toplevel()
            x = parent.winfo_rootx() + (parent.winfo_width() - top.winfo_width()) // 2
            y = parent.winfo_rooty() + (parent.winfo_height() - top.winfo_height()) // 2
            top.geometry(f"+{max(0, x)}+{max(0, y)}")
        except tk.TclError:
            pass

        outcome: dict = {"result": None, "error": None}

        def _worker():
            try:
                # Unregister all first, then re-register.
                smappservice_install.unregister_all()
                outcome["result"] = smappservice_install.register_all(
                    install_menubar=install_menubar)
            except Exception as e:  # noqa: BLE001
                outcome["error"] = e

        def _on_done():
            try:
                pb.stop()
            except tk.TclError:
                pass
            try:
                top.grab_release()
                top.destroy()
            except tk.TclError:
                pass
            if outcome["error"] is not None:
                e = outcome["error"]
                messagebox.showerror(
                    "Rebuild failed",
                    f"{type(e).__name__}: {e}\n\n"
                    f"Detailed output in ~/MailWarden/logs/smappservice_install.log")
            else:
                result = outcome["result"] or {}
                registered = ", ".join(result.get("registered", [])) or "(none)"
                failed = ", ".join(result.get("failed", [])) or "(none)"
                messagebox.showinfo(
                    "Background services restarted",
                    f"Registered: {registered}\nFailed: {failed}\n\n"
                    f"If System Settings opens for approval, toggle MailWarden ON "
                    f"in Login Items and Extensions.\n\n"
                    f"Detailed output in ~/MailWarden/logs/smappservice_install.log")
            try:
                self.app.diagnostics_tab.refresh()
            except Exception:
                pass

        def _poll_worker(thread):
            if thread.is_alive():
                self.after(150, lambda: _poll_worker(thread))
                return
            _on_done()

        t = threading.Thread(target=_worker, daemon=True)
        t.start()
        self.after(150, lambda: _poll_worker(t))

    # ------------------------------------------------------------------
    # Uninstall
    # ------------------------------------------------------------------
    @staticmethod
    def _derive_app_bundle() -> Path:
        """Best-effort path to the running MailWarden.app bundle.

        Inside the built bundle sys.executable is
        .../MailWarden.app/Contents/MacOS/python, so the bundle is three
        parents up (MacOS -> Contents -> MailWarden.app). If that doesn't
        resolve to a .app (e.g. running from a dev checkout), fall back to
        the standard install location.
        """
        try:
            candidate = Path(sys.executable).resolve().parent.parent.parent
            if candidate.suffix == ".app" and candidate.exists():
                return candidate
        except Exception:
            pass
        return Path("/Applications/MailWarden.app")

    @staticmethod
    def _stop_running_agents() -> None:
        """Best-effort: terminate any currently-running MailWarden agent
        processes so they don't linger until next login.

        We target them by bundle identifier prefix (com.mailwarden*) via
        NSRunningApplication and skip our own process. The filter/report
        agents are short-lived launchd-triggered runs and won't relaunch
        after unregister_all(); the menu-bar agent is the only persistent
        one, so terminating it here is the main win. If PyObjC isn't
        available or the lookup fails, we SKIP this step entirely and rely
        on the agents not relaunching after unregister (they clear at next
        login). We never kill our own Dashboard process here.
        """
        try:
            from AppKit import NSWorkspace, NSRunningApplication  # type: ignore
            import os as _os
            my_pid = _os.getpid()
            apps = NSWorkspace.sharedWorkspace().runningApplications()
            for app in apps:
                try:
                    bundle_id = app.bundleIdentifier()
                    if not bundle_id or not str(bundle_id).startswith("com.mailwarden"):
                        continue
                    if app.processIdentifier() == my_pid:
                        continue  # never terminate the Dashboard here
                    app.terminate()
                    startup_log.step(f"uninstall: terminated {bundle_id}")
                except Exception as e:  # noqa: BLE001
                    startup_log.step(f"uninstall: could not terminate agent: {e}")
        except Exception as e:  # noqa: BLE001
            # PyObjC unavailable or API changed — skip; unregister already
            # ensures they won't relaunch after the next login.
            startup_log.step(f"uninstall: skip stop-agents ({e})")

    @staticmethod
    def _trash_bundle(bundle: Path) -> tuple[bool, str | None]:
        """Move the app bundle to the Trash via AppKit. Returns (ok, error).

        Uses NSWorkspace.recycleURLs_completionHandler_, which presents
        macOS's own admin-password prompt if the bundle is root-owned (the
        .pkg installs it as root). Because the completion handler is async,
        we block briefly on a threading.Event to capture the result; this
        runs on a worker thread so the Tk UI stays responsive.
        """
        try:
            from Foundation import NSURL  # type: ignore
            from AppKit import NSWorkspace  # type: ignore
            import threading as _threading
            url = NSURL.fileURLWithPath_(str(bundle))
            done = _threading.Event()
            result: dict = {"ok": False, "err": None}

            def _handler(new_urls, error):
                if error is not None:
                    result["err"] = str(error)
                    result["ok"] = False
                else:
                    result["ok"] = True
                done.set()

            NSWorkspace.sharedWorkspace().recycleURLs_completionHandler_(
                [url], _handler)
            # The completion handler fires on a background dispatch queue, so
            # waiting here (off the Tk main thread) is safe. Generous timeout
            # covers the admin-password prompt the user may need to answer.
            if not done.wait(timeout=120):
                return False, "Timed out moving the app to the Trash."
            return result["ok"], result["err"]
        except Exception as e:  # noqa: BLE001
            return False, str(e)

    def _on_uninstall(self):
        """Gate the uninstall behind NATIVE messageboxes, not a custom
        Toplevel checkbox.

        A custom Toplevel checkbox can mis-render or land off-screen (that is
        how the data-loss incident's default-checked box went unnoticed).
        Native askyesno dialogs ALWAYS render and are centered by macOS, so the
        two destructive decisions are made there:

          (1) "Remove MailWarden?"  — app + background services. No=abort.
          (2) "Also PERMANENTLY DELETE your data…?" — defaults to No. This is a
              separate dialog so the data choice can never be made by accident
              alongside the app-removal choice.

        Even after a Yes on (2), _run_uninstall makes a verified backup (to a
        folder in your home directory) and asks ONE more time before anything
        is deleted.
        """
        parent = self.winfo_toplevel()

        if not messagebox.askyesno(
                "Remove MailWarden?",
                "This will remove MailWarden from your Mac:\n\n"
                "•  It turns off and removes MailWarden's background services "
                "(the spam filter, daily report, and menu bar).\n"
                "•  It moves the MailWarden app to the Trash.\n\n"
                "Your settings, history, and saved passwords are KEPT (in your "
                "home folder) unless you choose otherwise in the next step, so "
                "you can reinstall later without re-entering anything.\n\n"
                "Remove MailWarden now?",
                icon=messagebox.WARNING,
                default=messagebox.NO,
                parent=parent):
            return  # User declined the whole uninstall.

        # SEPARATE, second native dialog for the destructive data choice.
        # default=NO means hitting Return / closing the dialog KEEPS data.
        delete_data = messagebox.askyesno(
            "Also delete your data?",
            "Do you ALSO want to permanently delete your data from this Mac?\n\n"
            "This includes your Anthropic API key and your email account "
            "passwords, plus all settings, signal history, and logs.\n\n"
            "If you choose Yes, MailWarden first saves a verified backup of "
            "your data to a folder in your home directory and asks you to "
            "confirm once more before deleting anything.\n\n"
            "Choose No to keep your data (recommended).",
            icon=messagebox.WARNING,
            default=messagebox.NO,
            parent=parent)

        self._run_uninstall(bool(delete_data))

    def _backup_mailwarden_data(self):
        """Write a gzip tarball of ~/MailWarden to a sibling folder in the home
        directory root (~/MailWarden-uninstall-backup-<stamp>/), then REOPEN and
        VALIDATE it before reporting success. Returns (Path, None) on success
        or (None, error_message) on failure. Never raises — the caller treats
        any failure as 'do not delete'.

        The backup goes to the home folder root (NOT ~/Desktop and NOT inside
        ~/MailWarden) for two reasons: writing under the home root does not
        trigger the macOS Desktop/Documents TCC permission prompt, and a
        location outside ~/MailWarden survives the data delete that follows.

        Validation deliberately reopens the just-written archive (a
        write-then-read round-trip) and requires BOTH:
          •  member count > 0, and
          •  the config file (config/config.json) is present inside it.
        A backup that doesn't actually contain the config is worthless for the
        one thing the user most needs back (API key + email passwords), so we
        refuse to proceed with the delete unless the config is verifiably in
        the archive."""
        import tarfile
        from datetime import datetime as _dt
        try:
            stamp = _dt.now().strftime("%Y%m%d-%H%M%S")
            backup_dir = Path.home() / f"MailWarden-uninstall-backup-{stamp}"
            backup_dir.mkdir(parents=True, exist_ok=True)
            backup_path = backup_dir / f"MailWarden-backup-{stamp}.tar.gz"

            root_name = paths.MAILWARDEN_ROOT.name  # "MailWarden"
            with tarfile.open(backup_path, "w:gz") as tar:
                tar.add(paths.MAILWARDEN_ROOT, arcname=root_name)

            if not backup_path.exists() or backup_path.stat().st_size == 0:
                return None, "Backup file was not written or is empty."

            # Reopen the archive and validate its contents.
            # config.json lives at <root>/config/config.json; inside the tar it
            # is stored as "<root_name>/config/config.json".
            want_config = f"{root_name}/config/config.json"
            with tarfile.open(backup_path, "r:gz") as tar:
                names = tar.getnames()
            member_count = len(names)
            if member_count <= 0:
                return None, "Backup archive is empty (no files inside)."
            # Config presence: accept the exact path or any member ending with
            # config/config.json (robust to arcname edge cases).
            has_config = any(
                n == want_config or n.endswith("config/config.json")
                for n in names)
            if not has_config:
                return None, (
                    "Backup did not contain your config file "
                    "(config/config.json), so it cannot be trusted.")

            startup_log.step(
                f"uninstall: backup ok {backup_path} "
                f"size={backup_path.stat().st_size} "
                f"members={member_count} config_present=True")
            return backup_path, None
        except Exception as e:  # noqa: BLE001
            return None, f"{type(e).__name__}: {e}"

    def _run_uninstall(self, delete_data: bool):
        """Perform the uninstall on a worker thread, mirroring the
        Restart-all-background-services pattern: modal progress dialog, worker
        thread for the slow/destructive work, all Tk widget calls on the main
        thread via self.after().

        Each step is best-effort and logged so a later step still runs if an
        earlier one has trouble. Order: (a) unregister agents, (b) stop any
        running agent processes, (c) optionally delete user data, (d) move the
        bundle to Trash, (e) inform + quit.
        """
        import threading
        bundle = self._derive_app_bundle()
        startup_log.step(
            f"uninstall: start delete_data={delete_data} bundle={bundle}")

        # SAFETY GATE for the destructive data delete. Runs on the Tk main
        # thread BEFORE the worker starts, so messagebox/Tk calls are legal and
        # nothing is deleted unless (1) a home-folder backup is written and verified
        # and (2) the user confirms a second time. Any failure or decline here
        # downgrades to a data-preserving uninstall (delete_data=False) rather
        # than risking unrecoverable data loss.
        if delete_data:
            if paths.MAILWARDEN_ROOT.exists():
                backup_path, backup_err = self._backup_mailwarden_data()
                if backup_err is not None or backup_path is None:
                    startup_log.step(
                        f"uninstall: backup FAILED, aborting delete: "
                        f"{backup_err!r}")
                    messagebox.showerror(
                        "Backup failed — your data was NOT deleted",
                        "MailWarden could not create a backup of your data in "
                        "your home folder, so it did not delete anything.\n\n"
                        f"Details: {backup_err}\n\n"
                        "Nothing was removed. Please try again, or uninstall "
                        "without the delete-my-data option.")
                    return
                # Reveal the verified backup in Finder so the user can see
                # exactly where it landed before confirming the delete.
                try:
                    subprocess.Popen(["open", "-R", str(backup_path)])
                except Exception as e:  # noqa: BLE001
                    startup_log.step(f"uninstall: reveal-in-Finder failed: {e}")
                if not messagebox.askyesno(
                        "Permanently delete your data?",
                        "A verified backup was saved here (it's been revealed "
                        "in Finder):\n\n"
                        f"{backup_path}\n\n"
                        "This backup stays in your home folder and is NOT "
                        "deleted by the uninstall, so you can restore from it "
                        "later.\n\n"
                        "MailWarden will now PERMANENTLY DELETE your data from "
                        "this Mac. This includes your Anthropic API key and "
                        "your email account passwords, plus all settings, "
                        "signal history, and logs.\n\n"
                        "Delete this data now?",
                        icon=messagebox.WARNING,
                        default=messagebox.NO):
                    startup_log.step("uninstall: 2nd confirm declined; keeping data")
                    delete_data = False
            else:
                # Nothing to back up or delete; proceed as a plain uninstall.
                delete_data = False

        top = tk.Toplevel(self)
        top.title("Uninstalling MailWarden")
        top.transient(self.winfo_toplevel())
        top.grab_set()
        # Render-safety: force the modal to be visible, raised, and focused so
        # it can never sit hidden or unfocused while it owns the grab.
        try:
            top.deiconify()
            top.lift()
            top.focus_force()
        except tk.TclError:
            pass
        top.resizable(False, False)
        ttk.Label(top, padding=(20, 16), text=(
            "Removing MailWarden's background services"
            + (" and data" if delete_data else "")
            + ",\nthen moving the app to the Trash. This may take a moment\n"
            "and macOS may ask for your password to finish.")).pack()
        pb = ttk.Progressbar(top, mode="indeterminate", length=340)
        pb.pack(padx=20, pady=(0, 16))
        pb.start(80)

        top.update_idletasks()
        try:
            # Center on the parent, then CLAMP to the visible screen so the
            # dialog can never land off-screen (a custom Toplevel landing
            # off-screen is exactly the render failure native dialogs avoid).
            parent = self.winfo_toplevel()
            w = top.winfo_width()
            h = top.winfo_height()
            x = parent.winfo_rootx() + (parent.winfo_width() - w) // 2
            y = parent.winfo_rooty() + (parent.winfo_height() - h) // 2
            screen_w = top.winfo_screenwidth()
            screen_h = top.winfo_screenheight()
            x = max(0, min(x, screen_w - w))
            y = max(0, min(y, screen_h - h))
            top.geometry(f"+{x}+{y}")
        except tk.TclError:
            pass

        outcome: dict = {"trash_ok": False, "trash_err": None, "error": None}

        def _worker():
            try:
                # (a) Unregister the 3 background agents so they won't relaunch.
                try:
                    smappservice_install.unregister_all()
                    startup_log.step("uninstall: unregister_all() done")
                except Exception as e:  # noqa: BLE001
                    startup_log.step(f"uninstall: unregister_all error: {e}")

                # (b) Stop any running agent processes (best-effort, skippable).
                self._stop_running_agents()

                # (c) Optionally delete the user-data dir (config holds the API
                # key + email passwords; logs; memory; signal history).
                #
                # REVERSIBLE delete: first os.rename the root to a sibling
                # "MailWarden.deleting-<timestamp>" (atomic on the same
                # filesystem — the dir either is fully renamed or not at all),
                # and ONLY THEN remove that. If the process is interrupted
                # between the rename and the rmtree, the data still exists under
                # the .deleting- name and can be recovered. A verified backup
                # in the home folder also already exists at this point.
                if delete_data and paths.MAILWARDEN_ROOT.exists():
                    import shutil
                    from datetime import datetime as _dt
                    stamp = _dt.now().strftime("%Y%m%d-%H%M%S")
                    staging = paths.MAILWARDEN_ROOT.parent / (
                        f"{paths.MAILWARDEN_ROOT.name}.deleting-{stamp}")
                    try:
                        os.rename(paths.MAILWARDEN_ROOT, staging)
                        startup_log.step(
                            f"uninstall: renamed data root to {staging}")
                    except OSError as e:
                        # Rename failed (e.g. cross-device, perms). Fall back to
                        # an in-place best-effort rmtree rather than leaving the
                        # data half-removed; the verified backup still protects
                        # the user.
                        startup_log.step(
                            f"uninstall: rename failed ({e}); rmtree in place")
                        staging = paths.MAILWARDEN_ROOT
                    shutil.rmtree(staging, ignore_errors=True)
                    startup_log.step(
                        f"uninstall: rmtree({staging}) "
                        f"root_exists_after={paths.MAILWARDEN_ROOT.exists()} "
                        f"staging_exists_after={staging.exists()}")

                # (d) Move the app bundle to the Trash automatically.
                ok, err = self._trash_bundle(bundle)
                outcome["trash_ok"] = ok
                outcome["trash_err"] = err
                startup_log.step(f"uninstall: trash ok={ok} err={err!r}")
            except Exception as e:  # noqa: BLE001
                outcome["error"] = e
                startup_log.step(f"uninstall: worker EXCEPTION: {e}")

        def _quit_app():
            # Tear down the Tk mainloop so app_entrypoint's finally cleans up
            # and the process exits — the app must not run after its bundle is
            # trashed. Mirrors _on_close().
            try:
                self.app.quit()
            except Exception:
                pass
            try:
                self.app.destroy()
            except Exception:
                pass

        def _on_done():
            try:
                pb.stop()
            except tk.TclError:
                pass
            try:
                top.grab_release()
                top.destroy()
            except tk.TclError:
                pass

            if outcome["error"] is not None:
                # Unexpected failure — show a clear message, not a stack trace.
                messagebox.showerror(
                    "Uninstall problem",
                    "Something went wrong while uninstalling MailWarden. "
                    "Its background services and data may already be removed.\n\n"
                    "To finish, drag MailWarden from your Applications folder "
                    "to the Trash.\n\n"
                    "Details were saved to "
                    "~/MailWarden/logs/app_startup.log (if that folder still "
                    "exists).")
                _quit_app()
                return

            if outcome["trash_ok"]:
                messagebox.showinfo(
                    "MailWarden uninstalled",
                    "MailWarden has been uninstalled. The app has been moved "
                    "to your Trash and will now quit.")
            else:
                # Trash move failed (e.g. permission denied on a root-owned
                # bundle and the user declined the admin prompt). Agents (and
                # data, if requested) are already handled; tell the user how to
                # finish.
                messagebox.showinfo(
                    "Almost done",
                    "MailWarden's background services"
                    + (" and data" if delete_data else "")
                    + " have been removed. To finish, drag MailWarden from "
                    "your Applications folder to the Trash.\n\n"
                    "MailWarden will now quit.")
            _quit_app()

        def _poll_worker(thread):
            if thread.is_alive():
                self.after(150, lambda: _poll_worker(thread))
                return
            _on_done()

        t = threading.Thread(target=_worker, daemon=True)
        t.start()
        self.after(150, lambda: _poll_worker(t))

    def _on_pause_toggle(self):
        config = config_io.load_config()
        accounts = config.get("accounts", [])
        if not accounts:
            return
        ui = config.setdefault("ui", {})
        currently_paused = not any(a.get("enabled") for a in accounts)
        if currently_paused:
            # Restore per-position. The old version keyed on account name,
            # which collapsed duplicate/empty names into the same bucket
            # and lost one account's enabled state on resume. Using index
            # survives duplicates and un-named accounts.
            pre_list = ui.get("_pre_pause_enabled_list")
            if isinstance(pre_list, list) and len(pre_list) == len(accounts):
                for a, was_enabled in zip(accounts, pre_list):
                    a["enabled"] = bool(was_enabled)
            else:
                # Fall back to the name-keyed legacy map (older configs) or
                # default everything to enabled if no snapshot exists.
                legacy = ui.get("_pre_pause_enabled", {})
                for a in accounts:
                    a["enabled"] = bool(legacy.get(a.get("name", ""), True))
            ui.pop("_pre_pause_enabled_list", None)
            ui.pop("_pre_pause_enabled", None)
            ui["paused"] = False
        else:
            ui["_pre_pause_enabled_list"] = [bool(a.get("enabled")) for a in accounts]
            ui.pop("_pre_pause_enabled", None)  # clear any legacy key
            for a in accounts:
                a["enabled"] = False
            ui["paused"] = True
        config_io.save_config(config)
        self.app.refresh_all()

    def _on_share_signals(self):
        signals = config_io.load_signals()
        # Strip learner_notes
        if "signals" in signals and isinstance(signals["signals"], dict):
            signals["signals"]["learner_notes"] = (
                "Contributed via Dashboard. Generic description.")
            signals["signals"]["known_sending_infrastructure"] = [
                s for s in signals["signals"].get("known_sending_infrastructure", [])
                if "@" not in s and "box" not in s.lower()
            ]
        import tempfile
        fd, tmp = tempfile.mkstemp(prefix="mailwarden-signals-", suffix=".json")
        os.close(fd)
        with open(tmp, "w") as f:
            json.dump(signals, f, indent=2)
        # Open Finder and compose email
        subprocess.Popen(["open", "-R", tmp])
        mailto = (f"mailto:{help_content.FEEDBACK_EMAIL}"
                  f"?subject=MailWarden%20signal%20contribution"
                  f"&body=Attaching%20signals%20file%20at%20{tmp}")
        webbrowser.open(mailto)
        messagebox.showinfo("Signal contribution",
                             "A scrubbed copy of your signals was saved and "
                             "your email client is opening. Attach the file "
                             "shown in Finder to the email.")


# =============================================================================
# DIAGNOSTICS
# =============================================================================
# Parser helpers — keep in sync with payload/MailWarden/src/spam_filter.py.
# Inlined rather than imported so the Test Forward Parser shows what the
# CURRENT bundled code does, not whatever stale copy happens to be at
# ~/MailWarden/src/spam_filter.py. If the user's filter reports
# "could not parse" but this tool parses the same text cleanly, that's a
# smoking gun for a failed code refresh.
import re as _diag_re


def _diag_parse_from_address(header_value: str) -> dict:
    result = {"display_name": None, "address": None}
    if not header_value or not isinstance(header_value, str):
        return result
    # Some forwards through HTML intermediaries end up with &lt; &gt; in the
    # plain-text From: value. Decode entities up front so angle-bracket
    # matching works on the normal form.
    import html as _html_stdlib
    header_value = _html_stdlib.unescape(header_value.strip())
    if not header_value:
        return result
    # Decode RFC 2047 encoded-word if present (lightweight; the real
    # filter uses email.header.decode_header for this).
    try:
        import email.header
        parts = email.header.decode_header(header_value)
        decoded = []
        for part, charset in parts:
            if isinstance(part, bytes):
                decoded.append(part.decode(charset or "utf-8", errors="replace"))
            else:
                decoded.append(part)
        header_value = " ".join(decoded)
    except Exception:
        pass
    angle = _diag_re.search(r'^(.*?)<([^>]+@[^>]+)>\s*$', header_value)
    if angle:
        display_name = angle.group(1).strip().strip('"').strip("'").strip()
        addr = angle.group(2).strip().lower()
        if _diag_re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', addr):
            result["address"] = addr
            result["display_name"] = display_name if display_name else None
        return result
    bare = header_value.strip().strip('"').strip("'").strip()
    if _diag_re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', bare):
        result["address"] = bare.lower()
    return result


def _diag_html_to_text(html: str) -> str:
    if not html:
        return ""
    import html as _html_module
    text = _diag_re.sub(r'<\s*br\s*/?\s*>', '\n', html, flags=_diag_re.IGNORECASE)
    text = _diag_re.sub(r'<\s*/\s*(p|div|tr|li|h[1-6]|blockquote)\s*>',
                        '\n', text, flags=_diag_re.IGNORECASE)
    text = _diag_re.sub(r'<\s*(script|style)[^>]*>.*?<\s*/\s*\1\s*>',
                        '', text, flags=_diag_re.IGNORECASE | _diag_re.DOTALL)
    text = _diag_re.sub(r'<[^>]+>', '', text)
    try:
        text = _html_module.unescape(text)
    except Exception:
        pass
    return text.strip()


def _diag_parse_forwarded_email(plain_body: str, html_body: str = "") -> dict:
    result = {
        "user_explanation": "[No explanation provided]",
        "original_from": "",
        "original_subject": "",
        "original_date": "",
        "original_body": "",
        "_divider_kind": "none",
        "_source": "plain",
    }
    body = (plain_body or "").strip()
    if not body and html_body:
        body = _diag_html_to_text(html_body)
        result["_source"] = "html"
    if not body:
        return result

    lines = body.split("\n")
    divider_idx = None
    divider_kind = "none"
    for i, line in enumerate(lines):
        stripped = line.strip()
        # Strip leading quote chars before matching dividers — Apple Mail
        # quotes the whole forwarded block with "> " when the user forwards
        # something already in a thread. Mirror the same fix in
        # spam_filter.py:parse_forwarded_email.
        unquoted = _diag_re.sub(r'^(\s*>\s*)+', '', stripped).strip()
        if _diag_re.match(r'-{3,}.*[Ff]orward.*-{3,}', unquoted):
            divider_idx = i; divider_kind = "dashes+forward"; break
        if unquoted == "Begin forwarded message:":
            divider_idx = i; divider_kind = "apple-mail"; break
        if _diag_re.match(r'^-{3,}\s*[Oo]riginal\s+[Mm]essage\s*-{3,}\s*$', unquoted):
            divider_idx = i; divider_kind = "outlook"; break
        if _diag_re.match(r'^-{3,}\s*$', unquoted) and i + 1 < len(lines):
            nxt = _diag_re.sub(r'^(\s*>\s*)+', '', lines[i + 1].strip()).strip()
            if nxt.lower().startswith("from:"):
                divider_idx = i; divider_kind = "dashes+from-next"; break

    if divider_idx is not None:
        result["_divider_kind"] = divider_kind
        explanation = "\n".join(lines[:divider_idx]).strip()
        if explanation:
            result["user_explanation"] = explanation
        # Strip leading quote-prefix chars but NOT trailing/leading whitespace
        # here — RFC 2822 continuation lines (starting with space/tab) need
        # their leading whitespace preserved so the fold-joiner below works.
        below_lines = [_diag_re.sub(r'^(\s*>\s*)+', '', ln) for ln in lines[divider_idx + 1:]]
        below = "\n".join(below_lines)
        # Unfold RFC 2822 folded headers: a continuation line begins with
        # whitespace (space or tab) and logically belongs to the prior header's
        # value. Collapse the line break + leading whitespace into a single space.
        below = _diag_re.sub(r'\n[ \t]+', ' ', below)
        frm = _diag_re.search(r'(?im)^\s*from:\s*(.+)$', below)
        sub = _diag_re.search(r'(?im)^\s*subject:\s*(.+)$', below)
        dat = _diag_re.search(r'(?im)^\s*date:\s*(.+)$', below)
        if frm:
            result["original_from"] = frm.group(1).strip()
        if sub:
            result["original_subject"] = sub.group(1).strip()
        if dat:
            result["original_date"] = dat.group(1).strip()
        return result

    inline = _diag_re.search(
        r'On\s+[^\n]{3,120}?,\s*(.+?)\s*<([^>\s]+@[^>\s]+)>\s*wrote:',
        body)
    if inline:
        result["_divider_kind"] = "inline-quote-on-wrote"
        name = inline.group(1).strip().strip('"').strip("'").strip()
        addr = inline.group(2).strip()
        result["original_from"] = f'{name} <{addr}>' if name else addr
        exp = body[:inline.start()].strip()
        if exp:
            result["user_explanation"] = exp
        return result

    # Fallback: "On <date>, bare@address.com wrote:" — no angle brackets
    bare_inline = _diag_re.search(
        r'On\s+[^\n]{3,120}?,\s*([^<>\s]+@[^<>\s]+)\s+wrote:',
        body)
    if bare_inline:
        result["_divider_kind"] = "inline-quote-on-wrote-bare"
        result["original_from"] = bare_inline.group(1).strip()
        exp = body[:bare_inline.start()].strip()
        if exp:
            result["user_explanation"] = exp
        return result

    # Fallback 2a: wrapped-date inline attribution. Some iOS Mail locales put
    # the date across two lines. Bounded to 200 chars total to prevent runaway
    # matches across unrelated body paragraphs.
    # Three capture groups: (date_fragment, display_name, address).
    # The greedy .{3,200} date group backtracks to the last comma before the
    # display name, ensuring the name group captures only the actual name.
    wrapped = _diag_re.search(
        r'On\s+(.{3,200}),\s*(.+?)\s*<([^>\s]+@[^>\s]+)>\s*wrote:',
        body, _diag_re.DOTALL)
    if wrapped:
        result["_divider_kind"] = "inline-quote-wrapped-date"
        name = wrapped.group(2).strip().strip('"').strip("'").strip()
        if "\n" not in name and len(name) <= 80:
            addr = wrapped.group(3).strip()
            result["original_from"] = f'{name} <{addr}>' if name else addr
            exp = body[:wrapped.start()].strip()
            if exp:
                result["user_explanation"] = exp
            return result

    short = _diag_re.search(
        r'(.+?)\s*<([^>\s]+@[^>\s]+)>\s*wrote:\s*$',
        body, _diag_re.MULTILINE)
    if short:
        name = short.group(1).strip().strip('"').strip("'").strip()
        if name and "\n" not in name and len(name) <= 80:
            result["_divider_kind"] = "inline-quote-short"
            addr = short.group(2).strip()
            result["original_from"] = f'{name} <{addr}>'
            exp = body[:short.start()].strip()
            if exp:
                result["user_explanation"] = exp
            return result

    result["user_explanation"] = body[:1000]
    return result


class DiagnosticsTab(ttk.Frame):
    """Self-service diagnostic surface so users never need to paste terminal
    output to answer "why didn't the filter pick up my forward?". Three
    sections: build info (prove new code is installed), Test Forward Parser
    (paste a forward body, see exactly what the parser extracted), and
    Open Logs (one-click access to spam_filter.log / app_startup.log /
    smappservice_install.log)."""

    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        self._scroll = _ScrollableTab(self)
        self._scroll.pack(fill=tk.BOTH, expand=True)
        self._f = ttk.Frame(self._scroll.body, padding=(16, 12))
        self._f.pack(fill=tk.BOTH, expand=True)
        self._build()
        self.refresh()

    def _build(self):
        ttk.Label(self._f, text="Diagnostics", style="Heading.TLabel").pack(
            anchor=tk.W, pady=(0, 4))
        ttk.Label(self._f, style="Muted.TLabel",
                  text="What the filter actually sees. No terminal required.").pack(
            anchor=tk.W, pady=(0, 12))

        # --- Build info ---
        build_box = ttk.LabelFrame(self._f, text="Build info",
                                    padding=(12, 8))
        build_box.pack(fill=tk.X, pady=(0, 10))
        self._version_lbl = ttk.Label(build_box, text="—")
        self._version_lbl.pack(anchor=tk.W)
        self._installed_lbl = ttk.Label(build_box, text="—", style="Muted.TLabel")
        self._installed_lbl.pack(anchor=tk.W, pady=(2, 0))
        self._drift_lbl = ttk.Label(build_box, text="", style="Muted.TLabel",
                                     wraplength=700)
        self._drift_lbl.pack(anchor=tk.W, pady=(4, 0))

        btns = ttk.Frame(build_box)
        btns.pack(anchor=tk.W, pady=(8, 0))
        ttk.Button(btns, text="Open filter log",
                   command=self._open_filter_log).pack(side=tk.LEFT)
        ttk.Button(btns, text="Open app startup log",
                   command=self._open_startup_log).pack(side=tk.LEFT, padx=(6, 0))
        ttk.Button(btns, text="Open launchd install log",
                   command=self._open_launchd_log).pack(side=tk.LEFT, padx=(6, 0))

        # --- Menu bar status ---
        # Self-service diagnostic for the menu bar agent. Without this,
        # a missing icon could be any of: launchd never bootstrapped the
        # agent, the agent crashed on init, the .icns rendered at zero
        # size, or rumps fell back to text. Each row below answers one
        # of those. The "Reinstall menu bar agent" button is the
        # one-click rescue when the agent simply isn't loaded.
        menubar_box = ttk.LabelFrame(self._f, text="Menu bar status",
                                      padding=(12, 8))
        menubar_box.pack(fill=tk.X, pady=(0, 10))

        # Auto-launch at login indicator (Task 3) — placed first in this box
        self._autolaunch_lbl = ttk.Label(menubar_box, text="—",
                                          style="Muted.TLabel", wraplength=700)
        self._autolaunch_lbl.pack(anchor=tk.W)
        self._autolaunch_detail_lbl = ttk.Label(menubar_box, text="",
                                                  style="Muted.TLabel",
                                                  wraplength=700)
        self._autolaunch_detail_lbl.pack(anchor=tk.W, pady=(2, 8))

        self._mb_plist_lbl = ttk.Label(menubar_box, text="—",
                                        style="Muted.TLabel", wraplength=700)
        self._mb_plist_lbl.pack(anchor=tk.W)
        self._mb_loaded_lbl = ttk.Label(menubar_box, text="—",
                                         style="Muted.TLabel", wraplength=700)
        self._mb_loaded_lbl.pack(anchor=tk.W, pady=(2, 0))
        self._mb_icon_lbl = ttk.Label(menubar_box, text="—",
                                       style="Muted.TLabel", wraplength=700)
        self._mb_icon_lbl.pack(anchor=tk.W, pady=(2, 0))

        # Menu bar log excerpt
        ttk.Label(menubar_box, text="Menu bar log (start of file — "
                   "actual error appears here):",
                   style="Muted.TLabel").pack(anchor=tk.W, pady=(8, 2))
        log_wrap = ttk.Frame(menubar_box)
        log_wrap.pack(fill=tk.X)
        self._mb_log_text = tk.Text(log_wrap, height=8, wrap=tk.WORD,
                                     font=("Menlo", 10),
                                     background="#f5f5f5",
                                     foreground="#111111",
                                     insertbackground="#111111")
        mb_sb = ttk.Scrollbar(log_wrap, orient=tk.VERTICAL,
                               command=self._mb_log_text.yview)
        self._mb_log_text.configure(yscrollcommand=mb_sb.set)
        self._mb_log_text.pack(side=tk.LEFT, fill=tk.X, expand=True)
        mb_sb.pack(side=tk.RIGHT, fill=tk.Y)
        self._mb_log_text.config(state=tk.DISABLED)

        # smappservice_install.log last 15 lines (v1.6.0: replaces launchd_install.log)
        ttk.Label(menubar_box, text="smappservice_install.log (last 15 lines):",
                   style="Muted.TLabel").pack(anchor=tk.W, pady=(8, 2))
        ldlog_wrap = ttk.Frame(menubar_box)
        ldlog_wrap.pack(fill=tk.X)
        self._ld_log_text = tk.Text(ldlog_wrap, height=6, wrap=tk.WORD,
                                     font=("Menlo", 10),
                                     background="#f5f5f5",
                                     foreground="#111111",
                                     insertbackground="#111111")
        ld_sb = ttk.Scrollbar(ldlog_wrap, orient=tk.VERTICAL,
                               command=self._ld_log_text.yview)
        self._ld_log_text.configure(yscrollcommand=ld_sb.set)
        self._ld_log_text.pack(side=tk.LEFT, fill=tk.X, expand=True)
        ld_sb.pack(side=tk.RIGHT, fill=tk.Y)
        self._ld_log_text.config(state=tk.DISABLED)

        # Deployed plist content for all three agents
        ttk.Label(menubar_box, text="Deployed launchd plists:",
                   style="Muted.TLabel").pack(anchor=tk.W, pady=(8, 2))
        plist_wrap = ttk.Frame(menubar_box)
        plist_wrap.pack(fill=tk.X)
        self._mb_plist_text = tk.Text(plist_wrap, height=10, wrap=tk.NONE,
                                       font=("Menlo", 9),
                                       background="#f5f5f5",
                                       foreground="#111111",
                                       insertbackground="#111111")
        plist_sb_y = ttk.Scrollbar(plist_wrap, orient=tk.VERTICAL,
                                    command=self._mb_plist_text.yview)
        plist_sb_x = ttk.Scrollbar(plist_wrap, orient=tk.HORIZONTAL,
                                    command=self._mb_plist_text.xview)
        self._mb_plist_text.configure(yscrollcommand=plist_sb_y.set,
                                       xscrollcommand=plist_sb_x.set)
        plist_sb_y.pack(side=tk.RIGHT, fill=tk.Y)
        plist_sb_x.pack(side=tk.BOTTOM, fill=tk.X)
        self._mb_plist_text.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self._mb_plist_text.config(state=tk.DISABLED)

        mb_btns = ttk.Frame(menubar_box)
        mb_btns.pack(anchor=tk.W, pady=(8, 0))
        ttk.Button(mb_btns, text="Refresh status",
                   command=self._refresh_menubar_status).pack(side=tk.LEFT)
        ttk.Button(mb_btns, text="Reinstall menu bar agent",
                   command=self._on_reinstall_menubar).pack(
                       side=tk.LEFT, padx=(6, 0))
        # v1.6.0: deep-link to System Settings Login Items so users can
        # approve MailWarden without navigating there manually.
        ttk.Button(mb_btns, text="Open Login Items settings",
                   command=self._on_open_login_items).pack(
                       side=tk.LEFT, padx=(6, 0))
        ttk.Button(mb_btns, text="Open menu bar log",
                   command=self._open_menubar_log).pack(
                       side=tk.LEFT, padx=(6, 0))
        ttk.Button(mb_btns, text="Copy diagnostic to clipboard",
                   command=self._copy_menubar_diagnostic).pack(
                       side=tk.LEFT, padx=(6, 0))

        # --- Test Forward Parser ---
        test_box = ttk.LabelFrame(
            self._f,
            text="Test Forward Parser — paste the body of a forwarded email",
            padding=(12, 8))
        test_box.pack(fill=tk.BOTH, expand=True)
        ttk.Label(
            test_box, style="Muted.TLabel", wraplength=700,
            text=("Copy the body text of an email you forwarded as "
                  "\"Fwd: Whitelist\" or \"Fwd: Blacklist All\" into the box "
                  "below, then click Test. The tool shows what this build of "
                  "MailWarden would extract as the original sender. If it "
                  "works here but your filter still replies \"Could not parse\" "
                  "or \"already on the list,\" the filter is running an "
                  "older copy of the code — use Settings → Restart all "
                  "background services and reinstall the latest .pkg.")).pack(
                      anchor=tk.W, pady=(0, 8))
        # Fixed-height input + button row + fixed-height output — no `expand`
        # on either Text widget so one can never squeeze the other to zero.
        # Scrollbars keep the widgets usable when content exceeds the box.
        input_wrap = ttk.Frame(test_box)
        input_wrap.pack(fill=tk.X, pady=(0, 4))
        # Explicit fg/bg/cursor color — the app-wide ttk theme sets a light
        # default foreground that rendered as white-on-white inside tk.Text
        # widgets, making pasted content invisible until a mouse drag
        # selected it. Setting colors explicitly on Text overrides the
        # ttk default for this widget family.
        self._test_input = tk.Text(input_wrap, height=8, wrap=tk.WORD,
                                    font=("Menlo", 11),
                                    background="#ffffff",
                                    foreground="#111111",
                                    insertbackground="#111111")
        in_sb = ttk.Scrollbar(input_wrap, orient=tk.VERTICAL,
                               command=self._test_input.yview)
        self._test_input.configure(yscrollcommand=in_sb.set)
        self._test_input.pack(side=tk.LEFT, fill=tk.X, expand=True)
        in_sb.pack(side=tk.RIGHT, fill=tk.Y)

        row = ttk.Frame(test_box)
        row.pack(fill=tk.X, pady=(2, 6))
        ttk.Button(row, text="Test parser", style="Primary.TButton",
                   command=self._on_test_parser).pack(side=tk.LEFT)
        ttk.Button(row, text="Clear",
                   command=self._on_clear_parser).pack(
                       side=tk.LEFT, padx=(6, 0))

        ttk.Label(test_box, text="Result:", style="Muted.TLabel").pack(
            anchor=tk.W, pady=(4, 2))
        output_wrap = ttk.Frame(test_box)
        output_wrap.pack(fill=tk.BOTH, expand=True)
        self._test_output = tk.Text(output_wrap, height=12, wrap=tk.WORD,
                                     font=("Menlo", 11),
                                     background="#f5f5f5",
                                     foreground="#111111",
                                     insertbackground="#111111")
        out_sb = ttk.Scrollbar(output_wrap, orient=tk.VERTICAL,
                                command=self._test_output.yview)
        self._test_output.configure(yscrollcommand=out_sb.set)
        self._test_output.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        out_sb.pack(side=tk.RIGHT, fill=tk.Y)
        # Start with a visible placeholder so the area is obvious before the
        # first test run. Otherwise an empty output box looks broken.
        self._test_output.insert(
            tk.END,
            "Paste a forwarded email body above and click Test parser.\n"
            "Results appear here — divider kind, extracted From, "
            "parsed address + display name, and what the Fwd: handlers "
            "would do with this forward.")
        self._test_output.config(state=tk.DISABLED)

    def refresh(self):
        self._version_lbl.config(
            text=f"This build: MailWarden v{help_content.VERSION}")
        # Installed filter version from installer_state.json
        try:
            state = config_io.load_installer_state()
            inst = state.get("filter_version", "unknown")
        except Exception:
            inst = "unknown"
        self._installed_lbl.config(
            text=f"Installed runtime (~/MailWarden/src/): v{inst}")
        # Compare bundled code vs installed code via bootstrap's drift check.
        try:
            from . import bootstrap
            root = bootstrap._bundle_payload_root()
            drifted = bool(root and bootstrap._code_drifted(root))
        except Exception:
            drifted = False
        if drifted:
            self._drift_lbl.config(
                text=("⚠ Code drift detected: the bundle contains newer "
                      "filter code than ~/MailWarden/src/ has installed. "
                      "Quitting and relaunching MailWarden will refresh it "
                      "automatically; if that doesn't work, use Settings → "
                      "Restart all background services."),
                foreground="#a63")
        elif inst == help_content.VERSION:
            self._drift_lbl.config(
                text="✓ Bundled code and installed runtime match.",
                foreground="#262")
        else:
            self._drift_lbl.config(
                text="Versions differ but drift check could not run.",
                foreground="#555")
        self._refresh_menubar_status()

    def _refresh_menubar_status(self):
        """Populate the Menu bar status panel.

        Three independent checks because each one rules out a different
        failure class for the missing-icon symptom:
          1. Plist exists on disk → if not, install_menubar() never ran.
          2. launchctl print rc=0 → if not, launchd hasn't accepted it.
          3. icon file resolves and exists → if not, rumps falls back
             to text or shows nothing.
        Plus a tail of menubar.log so any rumps init crash is visible.
        """
        import os
        import subprocess as _sp

        # --- v1.6.0: SMAppService status for all three agents ---
        try:
            mb_status_int = smappservice_install.status_int_menubar()
            mb_status_text = smappservice_install.status_menubar()
        except Exception as e:
            mb_status_int = -1
            mb_status_text = f"Error reading status: {e}"

        if mb_status_int == 1:
            self._autolaunch_lbl.config(
                text="Auto-launch at login: \u2713 Enabled",
                foreground="#262")
            self._autolaunch_detail_lbl.config(
                text=("MailWarden's background services start automatically "
                      "at every login and persist across reboots. You don't "
                      "need to open the Dashboard to keep the filter running."),
                foreground="#555")
        elif mb_status_int == 2:
            self._autolaunch_lbl.config(
                text="Auto-launch at login: \u26a0 Approval pending",
                foreground="#a63")
            self._autolaunch_detail_lbl.config(
                text=("Open System Settings \u2192 General \u2192 Login Items "
                      "and Extensions, and toggle MailWarden ON. "
                      "Use the 'Open Login Items settings' button below."),
                foreground="#a63")
        elif mb_status_int == 3:
            self._autolaunch_lbl.config(
                text="Auto-launch at login: \u2717 Not found",
                foreground="#a33")
            self._autolaunch_detail_lbl.config(
                text="Installation incomplete. Reinstall MailWarden.pkg and relaunch.",
                foreground="#a33")
        else:
            self._autolaunch_lbl.config(
                text="Auto-launch at login: \u2717 Not registered",
                foreground="#a33")
            self._autolaunch_detail_lbl.config(
                text=("Background services are not registered. Go to Settings "
                      "\u2192 Background services and click \"Restart all "
                      "background services\" to re-enable."),
                foreground="#a33")

        # Show SMAppService status for menu bar agent
        self._mb_plist_lbl.config(
            text=f"Menu bar agent: {mb_status_text}",
            foreground="#262" if mb_status_int == 1 else "#a63" if mb_status_int == 2 else "#a33")

        # Show filter + report SMAppService status
        try:
            filter_status = smappservice_install.status_filter()
            report_status = smappservice_install.status_report()
        except Exception as e:
            filter_status = f"error: {e}"
            report_status = f"error: {e}"
        self._mb_loaded_lbl.config(
            text=f"Filter agent: {filter_status}\nReport agent: {report_status}",
            foreground="#262" if mb_status_int == 1 else "#555")

        try:
            from . import menu_bar as _mb
            icon_path, is_template = _mb._find_bundled_icon()
        except Exception as e:
            icon_path = None
            is_template = False
            self._mb_icon_lbl.config(
                text=f"? Icon lookup failed: {e}", foreground="#a83")
        if icon_path is not None:
            try:
                size_fmt = _fmt_file_size(icon_path.stat().st_size)
            except OSError:
                size_fmt = "?"
            template_note = " (template PNG — dark/light mode)" if is_template else ""
            self._mb_icon_lbl.config(
                text=(f"✓ Icon resolved: {icon_path} ({size_fmt}){template_note}"),
                foreground="#262")
        else:
            self._mb_icon_lbl.config(
                text=("✗ No icon found in any expected "
                      "location — the agent would fall back to "
                      "text \"MW\" in the menu bar."),
                foreground="#a33")

        log_path = paths.LOGS_DIR / "menubar.log"
        if log_path.exists():
            try:
                size = log_path.stat().st_size
                with log_path.open("rb") as f:
                    # Read the FIRST 8 KB. Python init crashes write
                    # their actual error message at the start of the
                    # log, then KeepAlive=true respawns the agent every
                    # 5 sec — so the LAST 4 KB is just a wall of repeat
                    # config dumps with the original error scrolled off.
                    # The first chunk has the real diagnostic.
                    head = f.read(8192).decode("utf-8", errors="replace")
                    if size > 8192:
                        head += (f"\n\n... [{size - 8192:,} more bytes — "
                                 f"if these are repeat crashes, click "
                                 f"\"Reinstall menu bar agent\" to clear "
                                 f"the log and capture a fresh trace] ...")
                    tail = head
            except OSError as e:
                tail = f"(could not read {log_path}: {e})"
            if not tail.strip():
                tail = ("(menubar.log exists but is empty — agent has "
                        "started cleanly with no stderr output)")
        else:
            tail = (f"(no log yet at {log_path} — the menu bar agent "
                    f"has not started, or has not written anything)")
        self._mb_log_text.config(state=tk.NORMAL)
        self._mb_log_text.delete("1.0", tk.END)
        self._mb_log_text.insert(tk.END, tail)
        self._mb_log_text.config(state=tk.DISABLED)

        # Populate smappservice_install.log last 15 lines
        # (v1.6.0: replaces launchd_install.log display)
        ld_log_path = paths.LOGS_DIR / "smappservice_install.log"
        if ld_log_path.exists():
            try:
                all_lines = ld_log_path.read_text(
                    encoding="utf-8", errors="replace").splitlines()
                ld_tail = "\n".join(all_lines[-15:])
                if not ld_tail.strip():
                    ld_tail = "(smappservice_install.log is empty)"
            except OSError as e:
                ld_tail = f"(could not read {ld_log_path}: {e})"
        else:
            ld_tail = f"(no SMAppService log yet at {ld_log_path})"
        self._ld_log_text.config(state=tk.NORMAL)
        self._ld_log_text.delete("1.0", tk.END)
        self._ld_log_text.insert(tk.END, ld_tail)
        self._ld_log_text.config(state=tk.DISABLED)

        # Populate bundled plist content (from inside the .app bundle)
        # v1.6.0: plists live in Contents/Library/LaunchAgents/, not
        # ~/Library/LaunchAgents/. Show the bundled copies.
        plist_sections = []
        import sys as _sys
        bundle_plist_dir = None
        here = __file__
        from pathlib import Path as _Path
        for ancestor in _Path(here).resolve().parents:
            candidate = ancestor / "Library" / "LaunchAgents"
            if candidate.is_dir():
                bundle_plist_dir = candidate
                break
        for label, plist_name in [
            ("filter", "com.mailwarden.filter.plist"),
            ("report", "com.mailwarden.report.plist"),
            ("menubar", "com.mailwarden.menubar.plist"),
        ]:
            plist_path = (bundle_plist_dir / plist_name) if bundle_plist_dir else None
            plist_sections.append(f"=== com.mailwarden.{label} ===")
            if plist_path is not None and plist_path.exists():
                try:
                    plist_sections.append(plist_path.read_text(
                        encoding="utf-8", errors="replace"))
                except OSError as e:
                    plist_sections.append(f"(could not read: {e})")
            else:
                plist_sections.append(
                    f"(not found in bundle: {plist_path or 'bundle dir not located'})")
            plist_sections.append("")
        plist_content = "\n".join(plist_sections)
        self._mb_plist_text.config(state=tk.NORMAL)
        self._mb_plist_text.delete("1.0", tk.END)
        self._mb_plist_text.insert(tk.END, plist_content)
        self._mb_plist_text.config(state=tk.DISABLED)

    def _open_menubar_log(self):
        """Open the full menubar.log in TextEdit."""
        log_path = paths.LOGS_DIR / "menubar.log"
        if log_path.exists():
            subprocess.Popen(
                ["open", str(log_path)],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL)
        else:
            messagebox.showinfo(
                "Menu bar log",
                f"No menu bar log found at {log_path}.\n\n"
                f"The menu bar agent has not started yet, or has not written "
                f"any output. Click Reinstall menu bar agent to try loading it.")

    def _copy_menubar_diagnostic(self):
        """Copy the full Menu bar status panel content to the clipboard."""
        parts = [
            self._autolaunch_lbl.cget("text"),
            self._autolaunch_detail_lbl.cget("text"),
            "",
            self._mb_plist_lbl.cget("text"),
            self._mb_loaded_lbl.cget("text"),
            self._mb_icon_lbl.cget("text"),
            "",
            "--- Menu bar log (start of file) ---",
            self._mb_log_text.get("1.0", tk.END).rstrip(),
            "",
            "--- smappservice_install.log (last 15 lines) ---",
            self._ld_log_text.get("1.0", tk.END).rstrip(),
            "",
            "--- Deployed plists ---",
            self._mb_plist_text.get("1.0", tk.END).rstrip(),
        ]
        text = "\n".join(parts)
        try:
            self.clipboard_clear()
            self.clipboard_append(text)
            self.update()
        except Exception as e:
            messagebox.showwarning(
                "Clipboard error",
                f"Could not copy to clipboard: {e}")

    def _on_reinstall_menubar(self):
        """One-click rescue: unregister then re-register the menu bar agent
        via SMAppService. Truncates menubar.log first so any new crash trace
        starts at byte 0 and is visible in the diagnostic panel.

        v1.6.0: uses SMAppService instead of launchctl bootout + bootstrap.
        """
        log_path = paths.LOGS_DIR / "menubar.log"
        try:
            if log_path.exists():
                log_path.write_text("", encoding="utf-8")
        except OSError:
            pass  # non-fatal
        try:
            smappservice_install.unregister_menubar()
            ok, err = smappservice_install.register_menubar()
        except Exception as e:
            messagebox.showerror(
                "Reinstall failed",
                f"Could not re-register the menu bar agent.\n\n"
                f"{type(e).__name__}: {e}\n\n"
                f"See ~/MailWarden/logs/smappservice_install.log for details.")
            self._refresh_menubar_status()
            return
        if not ok and err:
            messagebox.showerror(
                "Reinstall failed",
                f"SMAppService returned an error:\n{err}\n\n"
                f"See ~/MailWarden/logs/smappservice_install.log for details.")
        else:
            messagebox.showinfo(
                "Menu bar agent re-registered",
                "The menu bar agent has been re-registered via SMAppService. "
                "If System Settings opens for approval, toggle MailWarden ON "
                "in Login Items and Extensions.\n\n"
                "If the icon still doesn't appear, click Refresh status.")
        self._refresh_menubar_status()

    def _on_open_login_items(self):
        """Open the System Settings Login Items deep-link so the user can
        approve MailWarden without navigating there manually."""
        try:
            subprocess.Popen(
                ["open",
                 "x-apple.systempreferences:com.apple.LoginItems-Settings.extension"],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL)
        except OSError as e:
            messagebox.showerror("Could not open System Settings", str(e))

    # ---- Handlers ----

    def _open_filter_log(self):
        self._open_or_info(paths.FILTER_LOG,
                            "The filter hasn't run yet — no log to open.")

    def _open_startup_log(self):
        self._open_or_info(paths.LOGS_DIR / "app_startup.log",
                            "No startup log yet. Launch the app and come back.")

    def _open_launchd_log(self):
        # v1.6.0: smappservice_install.log replaces launchd_install.log
        self._open_or_info(paths.LOGS_DIR / "smappservice_install.log",
                            "No SMAppService install log yet — it's written the "
                            "first time MailWarden registers its background agents.")

    def _open_or_info(self, path: Path, empty_msg: str):
        if not path.exists():
            messagebox.showinfo("Not available yet", empty_msg)
            return
        try:
            subprocess.Popen(["open", str(path)])
        except OSError as e:
            messagebox.showerror("Could not open", str(e))

    def _on_clear_parser(self):
        """Clear both boxes. The previous version only cleared the input
        and left a stale result sitting in the output area, which made it
        look like the parser was stuck on the previous run."""
        self._test_input.delete("1.0", tk.END)
        self._set_output(
            "Paste a forwarded email body above and click Test parser.\n"
            "Results appear here — divider kind, extracted From, "
            "parsed address + display name, and what the Fwd: handlers "
            "would do with this forward.")

    def _set_output(self, text: str) -> None:
        """Always-show-something helper. Text widget is held disabled so the
        user can't edit it, but we briefly flip to NORMAL to write."""
        self._test_output.config(state=tk.NORMAL)
        self._test_output.delete("1.0", tk.END)
        self._test_output.insert(tk.END, text)
        self._test_output.config(state=tk.DISABLED)

    def _on_test_parser(self):
        # Bulletproof wrapper: if anything inside raises, show it to the user
        # instead of failing silently. That silent-fail mode was itself the
        # exact "terminal spot-check hell" we're trying to avoid.
        try:
            raw = self._test_input.get("1.0", tk.END)
            stripped = raw.strip()
            if not stripped:
                self._set_output(
                    "Nothing to parse. Paste the body of a forwarded email "
                    "(the visible text — with \"Begin forwarded message:\" or "
                    "similar — into the box above, then click Test parser.")
                return
            lower = stripped.lower()
            # Autodetect pasted HTML (webmail clients often put HTML on the
            # clipboard). If plain looks like HTML, route it through the
            # HTML-stripping fallback so the parser sees real text.
            looks_html = (
                "<html" in lower or "<body" in lower or "<div" in lower
                or "<br" in lower or "<p>" in lower or "<table" in lower
            )
            if looks_html:
                parsed = _diag_parse_forwarded_email("", stripped)
            else:
                parsed = _diag_parse_forwarded_email(stripped, "")
            sender = _diag_parse_from_address(parsed.get("original_from", ""))

            out = [
                f"Source body used: {parsed['_source']}"
                + ("  (pasted content looked like HTML — stripped before parsing)"
                   if looks_html else ""),
                f"Divider detected: {parsed['_divider_kind']}",
                "",
                f"Extracted From:    {parsed['original_from'] or '(empty)'}",
                f"Extracted Subject: {parsed['original_subject'] or '(empty)'}",
                f"Extracted Date:    {parsed['original_date'] or '(empty)'}",
                "",
                f"Parsed address:      {sender['address'] or '(none)'}",
                f"Parsed display name: {sender['display_name'] or '(none)'}",
                "",
            ]
            if not sender["address"] and not sender["display_name"]:
                out.append("→ Fwd: Whitelist would reply: 'Could Not Parse'")
                out.append("→ Fwd: Blacklist All would reply: 'Could Not Parse'")
                out.append("")
                out.append("The most common causes of this outcome:")
                out.append("  • Email client forwarded as an attachment.")
                out.append("  • Divider line missing from the copied body "
                           "(include the 'Begin forwarded message:' or "
                           "'-----Original Message-----' header).")
                out.append("  • Inline reply-quote ('On DATE, NAME wrote:') "
                           "was split across lines — try pasting more context.")
            else:
                if sender["address"]:
                    out.append(f"→ Fwd: Whitelist would ADD address "
                               f"{sender['address']} to whitelist.json")
                if sender["display_name"]:
                    out.append(f"→ Fwd: Blacklist All would ADD name "
                               f"{sender['display_name']!r} (subject to skip-list)")
                if sender["address"]:
                    out.append(f"→ Fwd: Blacklist Address would ADD "
                               f"{sender['address']} to blacklist.json")
            self._set_output("\n".join(out))
        except Exception as e:
            import traceback
            self._set_output(
                "The parser itself raised an error (please report):\n\n"
                f"{type(e).__name__}: {e}\n\n"
                + traceback.format_exc())


# =============================================================================
# HELP
# =============================================================================
class HelpTab(ttk.Frame):
    def __init__(self, parent, app: Dashboard):
        super().__init__(parent)
        self.app = app
        self._build()

    def _build(self):
        # Top button bar — EULA is always viewable here (the Installer EULA
        # click-through isn't easy to find again after install).
        btn_bar = ttk.Frame(self, padding=(16, 8))
        btn_bar.pack(side=tk.TOP, fill=tk.X)
        ttk.Button(btn_bar, text="View License Agreement (EULA)",
                   command=self._on_view_eula).pack(side=tk.LEFT)
        ttk.Button(btn_bar, text="Open GitHub repo",
                   command=lambda: __import__("webbrowser").open(
                       help_content.GITHUB_URL)).pack(side=tk.LEFT, padx=(8, 0))

        # Help text fills the remaining space with its own scrollbar.
        # Using the text widget's native scrollbar (not _ScrollableTab) here
        # because wrapping a fill/expand Text widget inside a canvas creates
        # a sizing conflict — the canvas never knows the text height.
        txt = tk.Text(self, wrap=tk.WORD,
                      background=theme.SURFACE, foreground=theme.TEXT,
                      relief=tk.FLAT, padx=18, pady=14,
                      font=theme.BODY_FONT,
                      spacing1=2, spacing2=4, spacing3=2,
                      borderwidth=0, highlightthickness=0)
        txt.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scroll = ttk.Scrollbar(self, command=txt.yview)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)
        txt.config(yscrollcommand=scroll.set)

        # Tags: readable paragraph spacing, generous heading lead-in.
        # lmargin1/2 align wrapped lines of a paragraph with the first line.
        txt.tag_configure("h1",
                          font=("Helvetica Neue", 20, "bold"),
                          foreground=theme.NAVY,
                          spacing1=0, spacing3=8)
        txt.tag_configure("h2",
                          font=("Helvetica Neue", 15, "bold"),
                          foreground=theme.NAVY,
                          spacing1=22, spacing3=6)
        txt.tag_configure("body",
                          font=theme.BODY_FONT,
                          spacing1=0, spacing2=3, spacing3=10,
                          lmargin1=0, lmargin2=0)
        txt.tag_configure("cmd",
                          font=("Menlo", 12, "bold"),
                          foreground=theme.NAVY,
                          spacing1=6, spacing3=2)
        txt.tag_configure("cmd_desc",
                          font=theme.BODY_FONT,
                          spacing3=8,
                          lmargin1=18, lmargin2=18)

        txt.insert(tk.END, "MailWarden Help\n", "h1")
        txt.insert(tk.END, help_content.HELP_TAB_INTRO + "\n", "body")

        for title, body in help_content.HELP_TAB_SECTIONS:
            txt.insert(tk.END, title + "\n", "h2")
            if body is None and title == "What each email command does":
                for cmd, desc in help_content.EMAIL_COMMAND_EXAMPLES:
                    txt.insert(tk.END, cmd + "\n", "cmd")
                    txt.insert(tk.END, desc + "\n", "cmd_desc")
            else:
                txt.insert(tk.END, body + "\n", "body")

        txt.config(state=tk.DISABLED)

    def _on_view_eula(self):
        from . import app_entrypoint
        eula_path = paths.MAILWARDEN_ROOT / "EULA.md"
        if not eula_path.exists():
            eula_path = app_entrypoint.get_bundled_defaults_dir() / "EULA.md"
        if not eula_path.exists():
            messagebox.showerror("EULA not found",
                                  f"Could not locate EULA.md at {eula_path}.")
            return
        EULAViewer(self.app, eula_path)


class EULAViewer(tk.Toplevel):
    """Read-only scrollable window showing the full EULA text."""

    def __init__(self, parent, eula_path):
        super().__init__(parent)
        self.title("MailWarden End User License Agreement")
        self.geometry("720x600")
        self.transient(parent)
        try:
            self.configure(bg=theme.BG)
        except Exception:
            pass

        frame = ttk.Frame(self, padding=(14, 12))
        frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(frame, text="End User License Agreement",
                  style="Heading.TLabel").pack(anchor=tk.W, pady=(0, 8))

        body = tk.Text(frame, wrap=tk.WORD,
                       background=theme.SURFACE, foreground=theme.TEXT,
                       relief=tk.FLAT, padx=12, pady=10,
                       font=theme.BODY_FONT)
        body.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scroll = ttk.Scrollbar(frame, command=body.yview)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)
        body.config(yscrollcommand=scroll.set)

        try:
            body.insert("1.0", eula_path.read_text(encoding="utf-8"))
        except (OSError, UnicodeDecodeError) as e:
            body.insert("1.0", f"(Could not load EULA text: {e})")
        body.config(state=tk.DISABLED)

        close_bar = ttk.Frame(self, padding=(14, 8))
        close_bar.pack(fill=tk.X)
        ttk.Button(close_bar, text="Close", command=self.destroy).pack(side=tk.RIGHT)


# =============================================================================
# helpers
# =============================================================================
def _last_filter_run() -> datetime | None:
    if not paths.FILTER_LOG.exists():
        return None
    return datetime.fromtimestamp(paths.FILTER_LOG.stat().st_mtime)


def _lock_active() -> bool:
    import time
    if not paths.FILTER_LOCK.exists():
        return False
    try:
        age = time.time() - paths.FILTER_LOCK.stat().st_mtime
        return age < 600
    except OSError:
        return False


def _determine_health(last_run: datetime | None) -> tuple[str, str]:
    if last_run is None:
        return "red", "No filter run on record yet — check back in 15 minutes."
    age = (datetime.now() - last_run).total_seconds()
    if age <= 20 * 60:
        return "green", f"All accounts healthy — last run {_fmt_age(age)} ago"
    if age <= 60 * 60:
        return "yellow", f"Filter is stale — last run {_fmt_age(age)} ago"
    return "red", f"Filter has not run recently — last run {_fmt_age(age)} ago"


def _fmt_age(sec: float) -> str:
    m = int(sec // 60)
    if m < 60:
        return f"{m}m"
    return f"{m // 60}h {m % 60}m"


def _decision_counts() -> tuple[int, int, int]:
    """Return (today, this_week, lifetime) evaluated-email counts from decisions.log.

    decisions.log is multi-line records separated by '  ---\\n'. Each record
    starts with '[YYYY-MM-DD HH:MM:SS] ACCOUNT: ...'. An earlier version of
    this function read line-by-line and tried line.split()[0] — which picked
    up '[2026-04-18' (with bracket) and failed to parse, so every count
    silently came back as zero. This version matches the daily_report.py
    parser: split on '  ---\\n', regex-extract the timestamp from each record.
    """
    import re
    if not paths.DECISIONS_LOG.exists():
        return 0, 0, 0
    today = datetime.now().date()
    week_start = today - timedelta(days=6)
    today_n = week_n = life_n = 0
    try:
        with paths.DECISIONS_LOG.open(encoding="utf-8", errors="replace") as f:
            content = f.read()
    except OSError:
        return 0, 0, 0
    ts_re = re.compile(r'\[(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})\]')
    for record in content.split("  ---\n"):
        record = record.strip()
        if not record:
            continue
        m = ts_re.search(record)
        if not m:
            continue
        try:
            d = datetime.strptime(m.group(1), "%Y-%m-%d %H:%M:%S").date()
        except ValueError:
            continue
        life_n += 1
        if d == today:
            today_n += 1
        if d >= week_start:
            week_n += 1
    return today_n, week_n, life_n


def _decision_counts_by_account() -> dict[str, int]:
    """Return a dict mapping account name -> number of decisions logged today.

    Uses the same record-splitting logic as _decision_counts so the two
    functions stay in sync.  Returns an empty dict (not a KeyError) for
    any account name not seen today — callers should .get(name, 0).
    """
    import re
    if not paths.DECISIONS_LOG.exists():
        return {}
    today = datetime.now().date()
    counts: dict[str, int] = {}
    try:
        with paths.DECISIONS_LOG.open(encoding="utf-8", errors="replace") as f:
            content = f.read()
    except OSError:
        return {}
    ts_re = re.compile(r'\[(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})\]')
    # The first line of each record is "[timestamp] ACCOUNT: name" —
    # ACCOUNT: is NOT at the start of any line, so use a plain search.
    acct_re = re.compile(r'\bACCOUNT: (.+)')
    for record in content.split("  ---\n"):
        record = record.strip()
        if not record:
            continue
        m_ts = ts_re.search(record)
        if not m_ts:
            continue
        try:
            d = datetime.strptime(m_ts.group(1), "%Y-%m-%d %H:%M:%S").date()
        except ValueError:
            continue
        if d != today:
            continue
        m_acct = acct_re.search(record)
        acct = m_acct.group(1).strip() if m_acct else "Unknown"
        counts[acct] = counts.get(acct, 0) + 1
    return counts


def _spam_killed_counts() -> tuple[int, int]:
    """Return (today, lifetime) counts of SPAM verdicts from decisions.log.

    A "spam caught" event is any decisions.log record whose DECISION line
    says SPAM. We count ALL SPAM verdicts — including dry-run entries — as
    "caught" because the classifier correctly identified the message as spam
    even when the move was suppressed by dry_run=true. The count therefore
    reflects classifier accuracy, not move actions. If you later need a
    "moved to junk" count exclusively, filter for ACTION lines that do NOT
    contain "DRY RUN".
    """
    import re
    if not paths.DECISIONS_LOG.exists():
        return 0, 0
    today = datetime.now().date()
    today_n = life_n = 0
    try:
        with paths.DECISIONS_LOG.open(encoding="utf-8", errors="replace") as f:
            content = f.read()
    except OSError:
        return 0, 0
    ts_re = re.compile(r'\[(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})\]')
    spam_re = re.compile(r'\bDECISION: SPAM\b')
    for record in content.split("  ---\n"):
        record = record.strip()
        if not record:
            continue
        if not spam_re.search(record):
            continue
        m = ts_re.search(record)
        if not m:
            continue
        try:
            d = datetime.strptime(m.group(1), "%Y-%m-%d %H:%M:%S").date()
        except ValueError:
            continue
        life_n += 1
        if d == today:
            today_n += 1
    return today_n, life_n


def _fmt_file_size(num_bytes: int) -> str:
    """Human-readable file size.

    Bytes below 1024 are shown as bytes (e.g. "168 B") to avoid the
    misleading "0 KB" that integer division produces for sub-kilobyte
    files like the menu-bar icon PNG.  KB with one decimal below 1 MB;
    MB with one decimal above.
    """
    if num_bytes < 1024:
        return f"{num_bytes} B"
    if num_bytes < 1024 * 1024:
        return f"{num_bytes / 1024:.1f} KB"
    return f"{num_bytes / 1024 / 1024:.1f} MB"


def _fetch_latest_release_tag() -> str | None:
    url = "https://api.github.com/repos/STR-Solutions-LLC/MailWarden/releases/latest"
    try:
        with urllib.request.urlopen(url, timeout=5) as resp:
            data = json.loads(resp.read())
        tag = data.get("tag_name", "")
        return tag.lstrip("v") if tag else None
    except Exception:
        return None


def _is_newer(latest: str, installed: str) -> bool:
    def parts(v: str) -> list[int]:
        out = []
        for p in v.split("."):
            try:
                out.append(int(p))
            except ValueError:
                out.append(0)
        return out
    return parts(latest) > parts(installed)


def _import_tabular_csv(path: Path) -> tuple[int, int]:
    added = skipped = 0
    wl = config_io.load_whitelist()
    bl = config_io.load_blacklist()
    with path.open(newline="") as f:
        reader = csv.reader(f)
        header = [h.strip().lower() for h in next(reader, [])]
        for row in reader:
            if not row:
                continue
            added_this, skipped_this = _absorb_row(header, row, wl, bl)
            added += added_this
            skipped += skipped_this
    config_io.save_whitelist(wl)
    config_io.save_blacklist(bl)
    return added, skipped


def _import_tabular_xlsx(path: Path) -> tuple[int, int]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl not installed; cannot import .xlsx")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, [])
    header = [str(h).strip().lower() if h is not None else "" for h in header_row]

    added = skipped = 0
    wl = config_io.load_whitelist()
    bl = config_io.load_blacklist()
    for row in rows_iter:
        if row is None:
            continue
        added_this, skipped_this = _absorb_row(
            header, [str(c) if c is not None else "" for c in row], wl, bl)
        added += added_this
        skipped += skipped_this
    config_io.save_whitelist(wl)
    config_io.save_blacklist(bl)
    return added, skipped


def _pick_cell(cells: dict, *names: str) -> str:
    """Return the first non-empty value for any of `names` in `cells`.

    `cells` has been lowercased+stripped during header parsing already,
    so we just try each alias in order. Lets the import accept common
    spreadsheet header variants ("List Type", "Category", "Email") without
    the user having to match the exact column name our code emits.
    """
    for n in names:
        v = cells.get(n)
        if v:
            return str(v).strip()
    return ""


def _absorb_row(header: list[str], row: list[str],
                wl: dict, bl: dict) -> tuple[int, int]:
    added = skipped = 0
    cells = dict(zip(header, row))
    list_ = _pick_cell(cells, "list", "list type", "type", "category",
                       "classification", "kind").lower()
    address = _pick_cell(cells, "address", "email", "email address",
                         "sender", "from")
    domain = _pick_cell(cells, "domain", "email domain", "sender domain")
    name = _pick_cell(cells, "display_name", "display name", "name",
                      "sender name", "from name")

    target_wl = list_ in ("w", "wl", "white", "whitelist", "allow", "allowlist",
                          "safe", "trusted")
    target_bl = list_ in ("b", "bl", "black", "blacklist", "block",
                          "blocklist", "banned", "deny")
    if not (target_wl or target_bl):
        # Infer: if there's a domain field, assume whitelist; else blacklist
        target_wl = bool(domain)
        target_bl = not target_wl

    if address:
        if target_wl:
            if address.lower() in {x.lower() for x in wl["addresses"]}:
                skipped += 1
            else:
                wl["addresses"].append(address); added += 1
        else:
            if address.lower() in {x.lower() for x in bl["addresses"]}:
                skipped += 1
            else:
                bl["addresses"].append(address); added += 1
    if domain:
        import re
        # Normalize domain: strip scheme, @-prefix, trailing dot, lowercase
        d = domain.strip()
        for scheme in ("https://", "http://"):
            if d.lower().startswith(scheme):
                d = d[len(scheme):].split("/")[0]
        if "@" in d:
            d = d.rsplit("@", 1)[1]
        d = d.lstrip("@").rstrip(".").lower()
        if d and "." in d and re.fullmatch(r"[a-z0-9.\-]+", d):
            if target_bl:
                if d.lower() in {x.lower() for x in bl.get("domains", [])}:
                    skipped += 1
                else:
                    bl.setdefault("domains", []).append(d); added += 1
            else:
                if d.lower() in {x.lower() for x in wl.get("domains", [])}:
                    skipped += 1
                else:
                    wl.setdefault("domains", []).append(d); added += 1
    if name and target_bl:
        if name.lower() in {x.lower() for x in bl["display_names"]}:
            skipped += 1
        else:
            bl["display_names"].append(name); added += 1
    return added, skipped


def _triple_confirm(parent, title: str, body: str) -> bool:
    prompts = [body,
               "Are you absolutely sure? This cannot be undone.",
               "Final confirmation — proceed?"]
    for p in prompts:
        if not messagebox.askyesno(title, p, parent=parent):
            return False
    return True


def run() -> int:
    from . import startup_log
    startup_log.step("Dashboard.run() entered")
    try:
        app = Dashboard()
        startup_log.step("Dashboard constructed; bringing to front")
        theme.bring_to_front(app)
        startup_log.step("Dashboard brought to front; entering mainloop")
        app.mainloop()
        startup_log.step("Dashboard mainloop exited")
    except BaseException as e:
        startup_log.fatal(e)
        raise
    return 0


if __name__ == "__main__":
    sys.exit(run())
